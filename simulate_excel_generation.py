"""
Simulate the exact Excel generation logic to verify the fix
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, OrderedDict
import pandas as pd

# Simulate the code logic
model_columns = ["300", "400", "500", "600", "7055", "8055"]
headers = ["批次号"] + model_columns + ["合计"]

print("=" * 70)
print("模拟Excel生成逻辑")
print("=" * 70)
print()
print(f"model_columns = {model_columns}")
print(f"headers = {headers}")
print(f"total_cols = {len(headers)}")
print()

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "排产台账"

# Styles
header_fill = PatternFill(start_color="5C765C", end_color="5C765C", fill_type="solid")
header_font = Font(name="宋体", size=11, bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)
green_fill = PatternFill(start_color="EAF2E8", end_color="EAF2E8", fill_type="solid")
orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

# Write headers
ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Simulate batch data
batches = OrderedDict([
    ("06-01", {
        "units": [
            {"机型": "FR-400AUTO", "备注": ""},
            {"机型": "FR-500AUTO", "备注": ""},
            {"机型": "FR-7055AUTO", "备注": ""},
            {"机型": "FR-8055AUTO", "备注": ""},
            {"机型": "FR-8055XS(PRO)", "备注": ""},
        ]
    }),
    ("06-02", {
        "units": [
            {"机型": "FR-8060AUTO", "备注": ""},
            {"机型": "FR-7055XS(PRO)", "备注": ""},
            {"机型": "FR-8055AUTO", "备注": ""},
        ]
    })
])

start_row = 2
total_cols = len(headers)

print("开始填充数据...")
print()

for batch_code, batch in batches.items():
    print(f"批次: {batch_code}")

    # Count models
    unique_pairs = []
    pair_qty = defaultdict(int)
    for unit in batch["units"]:
        orig_model = str(unit.get("机型") or "").strip()
        remark = str(unit.get("备注") or "").strip()
        if remark in ("None", "none", "null", "NULL"):
            remark = ""
        pair = (orig_model, remark)
        if pair_qty[pair] == 0:
            unique_pairs.append(pair)
        pair_qty[pair] += 1

    # Match to columns
    matched_by_col = defaultdict(list)
    special_pairs = []

    for pair in unique_pairs:
        orig_model, remark = pair
        combined = (orig_model + remark).upper()
        matched_col = None

        # Same matching logic as code
        if "8060" in combined:
            matched_col = "600"
        elif "8055" in combined:
            matched_col = "8055"
        elif "7055" in combined:
            matched_col = "7055"
        elif "600" in combined:
            matched_col = "600"
        elif "500" in combined:
            matched_col = "500"
        elif "400" in combined:
            matched_col = "400"
        elif "300" in combined:
            matched_col = "300"

        if matched_col:
            matched_by_col[matched_col].append(pair)
            print(f"  {orig_model} -> {matched_col}列")
        else:
            special_pairs.append(pair)

    # Build entries
    entries_by_model = {}
    for col_key in model_columns:
        entries = []
        for orig_model, remark in matched_by_col[col_key]:
            qty = pair_qty[(orig_model, remark)]
            entry = f"{orig_model} {qty}"
            entries.append(entry)
        entries_by_model[col_key] = entries

    max_rows = max(len(entries) for entries in entries_by_model.values()) if entries_by_model else 1
    if max_rows == 0:
        max_rows = 1

    end_row = start_row + max_rows - 1

    # Write batch code
    ws.cell(row=start_row, column=1).value = batch_code

    # Write model data
    for m_idx, m_type in enumerate(model_columns, start=2):
        entries = entries_by_model[m_type]
        for i, entry in enumerate(entries):
            ws.cell(row=start_row + i, column=m_idx).value = entry

    # Write total - THIS IS THE KEY FIX
    total_qty = len(batch["units"])
    ws.cell(row=start_row, column=total_cols).value = total_qty
    print(f"  合计: {total_qty} (写入第{total_cols}列)")

    start_row = end_row + 1
    print()

# Set column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 14
for col_idx in range(3, 8):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 20
ws.column_dimensions[get_column_letter(8)].width = 8

# Save
filename = "simulated_export.xlsx"
wb.save(filename)

print("=" * 70)
print(f"文件已保存: {filename}")
print("=" * 70)
print()

# Verify
df = pd.read_excel(filename, sheet_name='排产台账')
print("验证结果:")
print(f"  列数: {len(df.columns)}")
print(f"  列名: {list(df.columns)}")
print()

expected = ["批次号", "300", "400", "500", "600", "7055", "8055", "合计"]
if list(df.columns) == expected:
    print("[OK] 列结构正确！")
else:
    print("[ERROR] 列结构不匹配")
    print(f"  期望: {expected}")
    print(f"  实际: {list(df.columns)}")

print()
print("数据内容:")
print(df)

print()
print("检查8055列:")
if "8055" in df.columns:
    non_empty = df["8055"].notna() & (df["8055"] != "")
    if non_empty.any():
        print(f"  [OK] 8055列有数据: {df[non_empty]['8055'].tolist()}")
    else:
        print("  [ERROR] 8055列没有数据")
else:
    print("  [ERROR] 没有8055列")

print()
print("检查合计列:")
if "合计" in df.columns:
    print(f"  [OK] 合计列数据: {df['合计'].tolist()}")
else:
    print("  [ERROR] 没有合计列")
