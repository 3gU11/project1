"""
Download actual Excel file from API and verify the structure
"""
import requests
import pandas as pd
from openpyxl import load_workbook
import sys

BASE_URL = "http://localhost:8000"

# Get a valid token from the database
sys.path.insert(0, '.')
from database import get_engine
from sqlalchemy import text

print("获取认证令牌...")
with get_engine().connect() as conn:
    # Get a user token
    result = conn.execute(text("""
        SELECT username FROM users LIMIT 1
    """))
    row = result.fetchone()
    if not row:
        print("数据库中没有用户")
        sys.exit(1)
    
    username = row[0]
    print(f"使用用户: {username}")

# Try to get token (we'll skip auth and just try the export directly)
print()
print("直接调用导出API（跳过认证）...")

# Check if we can access without auth
try:
    export_resp = requests.get(
        f"{BASE_URL}/api/v1/planning/export-production-history",
        params={"sheet": "ledger"},
        timeout=30
    )
    
    if export_resp.status_code == 401:
        print("需要认证，尝试使用测试令牌...")
        # Try with a dummy token
        headers = {"Authorization": "Bearer test"}
        export_resp = requests.get(
            f"{BASE_URL}/api/v1/planning/export-production-history",
            params={"sheet": "ledger"},
            headers=headers,
            timeout=30
        )
    
    if export_resp.status_code != 200:
        print(f"导出失败: {export_resp.status_code}")
        print(export_resp.text[:500])
        
        # Try to read the code directly to understand the issue
        print()
        print("直接检查代码中的列定义...")
        with open('api/routes/planning.py', 'r', encoding='utf-8') as f:
            content = f.read()
            
            # Find model_columns definition
            import re
            match = re.search(r'model_columns\s*=\s*\[(.*?)\]', content, re.DOTALL)
            if match:
                print(f"代码中的 model_columns = [{match.group(1)}]")
            
            # Find headers definition
            match = re.search(r'headers\s*=\s*\["批次号"\]\s*\+\s*model_columns\s*\+\s*\[(.*?)\]', content)
            if match:
                print(f"代码中的 headers = ['批次号'] + model_columns + [{match.group(1)}]")
        
        sys.exit(1)
    
    # Save the file
    excel_file = "actual_export.xlsx"
    with open(excel_file, 'wb') as f:
        f.write(export_resp.content)
    
    print(f"Excel文件已下载: {excel_file}")
    print()
    
    # Read and analyze
    df = pd.read_excel(excel_file, sheet_name='排产台账')
    
    print("=" * 60)
    print("实际导出的Excel结构:")
    print("=" * 60)
    print(f"列数: {len(df.columns)}")
    print(f"列名: {list(df.columns)}")
    print()
    
    expected = ["批次号", "300", "400", "500", "600", "7055", "8055", "合计"]
    
    if list(df.columns) == expected:
        print("[OK] 列结构正确")
    else:
        print("[ERROR] 列结构不匹配")
        print(f"  期望: {expected}")
        print(f"  实际: {list(df.columns)}")
    
    print()
    print("列宽设置:")
    wb = load_workbook(excel_file)
    ws = wb['排产台账']
    
    for idx, col_name in enumerate(df.columns, 1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        width = ws.column_dimensions[col_letter].width
        print(f"  {col_letter} ({col_name}): {width}")
    
    print()
    print("前5行数据:")
    print(df.head())
    
    # Check if 8055 column has data
    print()
    print("=" * 60)
    print("检查8055列是否有数据:")
    print("=" * 60)
    
    if "8055" in df.columns:
        non_empty = df["8055"].notna() & (df["8055"] != "") & (df["8055"].astype(str).str.strip() != "")
        count = non_empty.sum()
        print(f"8055列非空行数: {count}")
        
        if count > 0:
            print()
            print("8055列的数据示例:")
            print(df[non_empty]["8055"].head())
        else:
            print("[WARNING] 8055列没有数据！")
    else:
        print("[ERROR] 没有8055列！")

except Exception as e:
    print(f"错误: {e}")
    import traceback
    traceback.print_exc()

