from typing import List, Dict, Any
from urllib.parse import unquote
import base64
import re
import asyncio
import hashlib

import os
import uuid
from datetime import datetime
import pandas as pd
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Query, Request, BackgroundTasks
import httpx
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
import json
from sqlalchemy import bindparam, text

from config import BASE_DIR, GO_SANDBOX_URL, GO_INTERNAL_TOKEN
from core.file_manager import delete_contract_file, save_contract_file
from crud.audit_logs import append_audit_log
from crud.contracts import get_contract_files
from crud.inventory import get_data, save_data
from crud.logs import append_log
from crud.model_dictionary import is_model_enabled
from crud.planning import get_factory_plan, get_factory_plan_v2, save_factory_plan
from crud.orders import allocate_inventory, get_orders, revert_to_inbound, save_orders
from api.routes.auth import get_current_operator_name, get_current_user_context, get_current_user_token
from utils.parsers import parse_alloc_dict
from database import get_engine

router = APIRouter(dependencies=[Depends(get_current_user_token)])

RUSH_AUTO_INSERT_ON_ENTRY = os.getenv("RUSH_AUTO_INSERT_ON_ENTRY", "").strip().lower() in {"1", "true", "yes", "on"}


def _parse_order_need_total(row: pd.Series) -> int:
    raw = str(row.get("需求机型", "") or "")
    total = 0
    for token_raw in re.split(r"[;；/,，]", raw):
        token = token_raw.strip()
        if not token:
            continue
        # 兼容 x3 / ×3 / :3
        m = re.search(r"(?:[x×:：]\s*)(\d+)\s*$", token, flags=re.IGNORECASE)
        if m:
            total += int(m.group(1))
    if total > 0:
        return total
    try:
        fallback = int(row.get("需求数量", 0) or 0)
    except Exception:
        fallback = 0
    return max(0, fallback)


def _parse_order_demand_counts(row: pd.Series) -> dict[str, int]:
    raw = str(row.get("需求机型", "") or "")
    counts: dict[str, int] = {}
    for token_raw in raw.split(";"):
        token = token_raw.strip()
        if not token:
            continue
        m = re.search(r"(?:[x×:：]\s*)(\d+)\s*$", token, flags=re.IGNORECASE)
        qty = int(m.group(1)) if m else 0
        model = re.sub(r"(?:[x×:：]\s*)\d+\s*$", "", token, flags=re.IGNORECASE).strip()
        model = model.replace("(加高)", "").strip()
        if model and qty > 0:
            counts[model] = counts.get(model, 0) + qty
    if counts:
        return counts
    fallback_model = raw.strip()
    fallback_qty = _parse_order_need_total(row)
    if fallback_model and fallback_qty > 0:
        return {fallback_model: fallback_qty}
    return {}


def _normalize_alloc_model(value: object) -> str:
    text_value = str(value or "").strip()
    if not text_value:
        return ""
    return text_value.replace("(加高)", "").replace("（加高）", "").strip()


def _is_high_model_hint(*values: object) -> bool:
    return any("加高" in str(value or "") for value in values)


def _machine_matches_model_requirement(row: pd.Series, required_model: str, required_high: bool = False) -> bool:
    row_model = _normalize_alloc_model(row.get("机型", ""))
    if row_model != _normalize_alloc_model(required_model):
        return False
    row_high = _is_high_model_hint(row.get("机型", ""), row.get("批次号", ""), row.get("合同备注", ""))
    if not required_high:
        return not row_high
    return row_high


def _inventory_rows_satisfy_order(order_row: pd.Series, rows: pd.DataFrame) -> bool:
    if rows.empty:
        return False
    demand_counts = _parse_order_demand_counts(order_row)
    if demand_counts:
        allocated_counts: dict[str, int] = {}
        for _, inv_row in rows.iterrows():
            model = _normalize_alloc_model(inv_row.get("机型", ""))
            if model:
                allocated_counts[model] = allocated_counts.get(model, 0) + 1
        for model, need in demand_counts.items():
            if allocated_counts.get(_normalize_alloc_model(model), 0) < need:
                return False
        return True

    need = _parse_order_need_total(order_row)
    return need > 0 and len(rows) >= need


def _extract_models_from_demand_text(raw_text: str) -> list[str]:
    models: list[str] = []
    raw = str(raw_text or "")
    for token_raw in raw.split(";"):
        token = token_raw.strip()
        if not token:
            continue
        model_part = re.sub(r"(?:[x×:：]\s*)\d+\s*$", "", token, flags=re.IGNORECASE).strip()
        model_name = model_part.replace("(加高)", "").strip()
        if model_name:
            models.append(model_name)
    return models


def _assert_models_in_dictionary(models: list[str]) -> None:
    invalid = [m for m in models if m and not is_model_enabled(m)]
    if invalid:
        unique_invalid = []
        seen = set()
        for item in invalid:
            if item not in seen:
                seen.add(item)
                unique_invalid.append(item)
        raise HTTPException(status_code=422, detail=f"机型不在字典中或未启用: {'，'.join(unique_invalid)}")


def _reconcile_completed_orders(df_orders: pd.DataFrame) -> pd.DataFrame:
    if df_orders.empty:
        return df_orders
    inv_df = get_data()
    if inv_df.empty:
        return df_orders

    for col in ["机型", "状态", "占用订单号", "流水号"]:
        if col not in inv_df.columns:
            inv_df[col] = ""
    inv_df["占用订单号"] = inv_df["占用订单号"].astype(str).str.strip()
    inv_df["状态"] = inv_df["状态"].astype(str).str.strip()

    changed = False
    for idx, row in df_orders.iterrows():
        oid = str(row.get("订单号", "") or "").strip()
        if not oid:
            continue
        status = str(row.get("status", "active") or "active")
        if status in ("deleted", "done"):
            continue

        shipped_rows = inv_df[(inv_df["占用订单号"] == oid) & (inv_df["状态"] == "已出库")]
        if _inventory_rows_satisfy_order(row, shipped_rows):
            df_orders.at[idx, "status"] = "done"
            changed = True
            continue

        pending_ship_rows = inv_df[(inv_df["占用订单号"] == oid) & (inv_df["状态"] == "待发货")]
        if status != "ready" and _inventory_rows_satisfy_order(row, pending_ship_rows):
            df_orders.at[idx, "status"] = "ready"
            changed = True

    if changed:
        save_orders(df_orders)
    return df_orders


class ContractItem(BaseModel):
    机型: str
    排产数量: int = Field(gt=0)
    备注: str = ""


class ContractEditPayload(BaseModel):
    客户名: str
    代理商: str
    要求交期: str
    items: List[ContractItem]
    confirmed_impact: bool = False
    mapping_decision: Dict[str, Any] | None = None


class StatusPayload(BaseModel):
    status: str


class PlanRowPayload(BaseModel):
    row_index: int
    allocation: Dict[str, int] = Field(default_factory=dict)


class PlanSavePayload(BaseModel):
    rows: List[PlanRowPayload]
    mark_to_planned: bool = True


class SalesOrderCreatePayload(BaseModel):
    客户名: str
    代理商: str = ""
    需求机型: str
    需求数量: int = Field(gt=0)
    备注: str = ""
    包装选项: str = ""
    发货时间: str = ""
    contract_ids: List[str] = Field(default_factory=list)


class SalesOrderUpdatePayload(BaseModel):
    客户名: str | None = None
    代理商: str | None = None
    需求机型: str | None = None
    需求数量: int | None = None
    备注: str | None = None
    包装选项: str | None = None
    发货时间: str | None = None
    status: str | None = None


class OrderAllocatePayload(BaseModel):
    selected_serial_nos: List[str] = Field(default_factory=list)


class OrderReleasePayload(BaseModel):
    selected_serial_nos: List[str] = Field(default_factory=list)
    all: bool = False


class BatchContractRowPayload(BaseModel):
    合同号: str
    客户名: str
    代理商: str = ""
    机型: str
    排产数量: int = Field(gt=0)
    要求交期: str
    备注: str = ""


class BatchContractCreatePayload(BaseModel):
    rows: List[BatchContractRowPayload]
    is_rush: bool = False
    save_mode: str = "sandbox"  # sandbox | spot
    dealer_order_no: str = ""  # 来源经销商订单号，创建成功后回写contract_no


class LinkOrderPayload(BaseModel):
    order_id: str


def _insert_production_queue(rows: list[dict[str, Any]]) -> int:
    """新合同录入后同步写入 production_queue，走排产队列调度。"""
    if not rows:
        return 0
    insert_values: list[dict[str, Any]] = []
    for row in rows:
        qty = int(row.get("排产数量") or 0)
        if qty <= 0:
            continue
        insert_values.append({
            "model_type": str(row.get("机型") or "").strip(),
            "contract_no": str(row.get("合同号") or "").strip(),
            "customer": str(row.get("客户名") or "").strip(),
            "dealer": str(row.get("代理商") or "").strip(),
            "due_date": str(row.get("要求交期") or "").strip(),
            "quantity_remaining": qty,
        })
    if not insert_values:
        return 0
    with get_engine().begin() as conn:
        for v in insert_values:
            result = conn.execute(
                text(
                    "UPDATE production_queue SET quantity_remaining = quantity_remaining + :qty "
                    "WHERE contract_no = :cid AND model_type = :model AND status = 'Waiting'"
                ),
                {"qty": v["quantity_remaining"], "cid": v["contract_no"], "model": v["model_type"]},
            )
            if result.rowcount == 0:
                conn.execute(
                    text(
                        "INSERT INTO production_queue (model_type, contract_no, customer, dealer, due_date, quantity_remaining, status) "
                        "VALUES (:model_type, :contract_no, :customer, :dealer, :due_date, :quantity_remaining, 'Waiting')"
                    ),
                    v,
                )
    return len(insert_values)


def _trigger_sandbox_recompute_sync(user_ctx: dict) -> bool:
    headers = {
        "Content-Type": "application/json",
        "X-Username": str(user_ctx.get("username") or ""),
        "X-Role": "Admin",
        "X-Original-Role": str(user_ctx.get("role") or ""),
        "X-User-ID": str(user_ctx.get("username") or ""),
    }
    if GO_INTERNAL_TOKEN:
        headers["X-Internal-Token"] = GO_INTERNAL_TOKEN
    
    try:
        with httpx.Client(timeout=120.0, trust_env=False) as client:
            resp = client.post(f"{GO_SANDBOX_URL}/api/forecast/recompute", headers=headers, json={"target_slot_no": 1, "is_clicked": False})
            resp.raise_for_status()
            return True
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"Auto-recompute failed: {e}")
        return False


def _auto_insert_rush_orders(rows: list[dict[str, Any]], user_ctx: dict, current_operator: str = "") -> int:
    if not rows:
        return 0

    headers = {
        "Content-Type": "application/json",
        "X-Username": str(user_ctx.get("username") or current_operator or ""),
        "X-Role": str(user_ctx.get("role") or ""),
        "X-User-ID": str(user_ctx.get("username") or current_operator or ""),
    }
    if GO_INTERNAL_TOKEN:
        headers["X-Internal-Token"] = GO_INTERNAL_TOKEN

    inserted = 0
    import logging
    logger = logging.getLogger(__name__)
    with httpx.Client(timeout=20.0, trust_env=False) as client:
        for row in rows:
            qty = int(row.get("qty") or 0)
            # 从 指定批次/来源 字典中取第一个批次号作为优先插入批次
            source_alloc = row.get("source_alloc") or {}
            preferred_batch_no = ""
            if isinstance(source_alloc, dict) and source_alloc:
                preferred_batch_no = str(next(iter(source_alloc), "")).strip()
            for _ in range(max(0, qty)):
                payload = {
                    "mode": "auto",
                    "rush_order": {
                        "contract_no": str(row.get("contract_no") or "").strip(),
                        "customer": str(row.get("customer") or "").strip(),
                        "model_type": str(row.get("model_type") or "").strip(),
                        "dealer_name": str(row.get("dealer_name") or "").strip(),
                        "due_date": str(row.get("due_date") or "").strip(),
                        "remark": str(row.get("remark") or "").strip(),
                        "preferred_batch_no": preferred_batch_no,
                    },
                    "reason": "急单录入后自动进入沙盘",
                }
                try:
                    resp = client.post(f"{GO_SANDBOX_URL}/api/units/rush-insert", headers=headers, json=payload)
                    if resp.status_code >= 400:
                        logger.warning("Rush auto insert failed for %s: %s", payload["rush_order"]["contract_no"], resp.text)
                        continue
                    with get_engine().begin() as conn:
                        conn.execute(text("""
                            UPDATE rush_order_queue
                            SET `status` = 'inserted', `updated_by` = :updated_by
                            WHERE `contract_no` = :contract_no
                              AND `model_type` = :model_type
                              AND COALESCE(`remark`, '') = :remark
                              AND `status` = 'pending'
                            ORDER BY id ASC
                            LIMIT 1
                        """), {
                            "updated_by": str(user_ctx.get("username") or current_operator or ""),
                            "contract_no": payload["rush_order"]["contract_no"],
                            "model_type": payload["rush_order"]["model_type"],
                            "remark": payload["rush_order"]["remark"],
                        })
                    inserted += 1
                except Exception as e:
                    logger.warning("Rush auto insert exception for %s: %s", payload["rush_order"]["contract_no"], e)
                    continue
    return inserted


def _insert_rush_order_queue(rows: list[dict[str, Any]], created_by: str = "") -> int:
    if not rows:
        return 0
    insert_rows: list[dict[str, Any]] = []
    for row in rows:
        qty = int(row.get("qty") or 0)
        if qty <= 0:
            continue
        contract_no = str(row.get("contract_no") or "").strip()
        model_type = str(row.get("model_type") or "").strip()
        if not contract_no or not model_type:
            continue
        due_date = str(row.get("due_date") or "").strip() or None
        for _ in range(qty):
            insert_rows.append({
                "contract_no": contract_no,
                "customer": str(row.get("customer") or "").strip(),
                "dealer_name": str(row.get("dealer_name") or "").strip(),
                "model_type": model_type,
                "due_date": due_date,
                "remark": str(row.get("remark") or "").strip(),
                "source": "contract",
                "status": "pending",
                "created_by": str(created_by or "").strip(),
                "updated_by": str(created_by or "").strip(),
            })
    if not insert_rows:
        return 0
    with get_engine().begin() as conn:
        conn.execute(
            text("""
                INSERT INTO rush_order_queue
                    (contract_no, customer, dealer_name, model_type, due_date, remark, source, status, created_by, updated_by)
                VALUES
                    (:contract_no, :customer, :dealer_name, :model_type, :due_date, :remark, :source, :status, :created_by, :updated_by)
            """),
            insert_rows,
        )
    return len(insert_rows)


def _clean_contract_ids(values: List[str] | None) -> list[str]:
    if isinstance(values, str):
        values = [v for v in re.split(r"[,\s;，；]+", values) if v]
    result: list[str] = []
    seen = set()
    for item in values or []:
        cid = str(item or "").strip()
        if cid and cid not in seen:
            seen.add(cid)
            result.append(cid)
    return result


def _user_id_from_context(current_user) -> str:
    return current_user.get("username") if isinstance(current_user, dict) else ""


def _occupy_inventory_for_order(contract_ids: list[str], order_id: str) -> int:
    contract_ids = _clean_contract_ids(contract_ids)
    order_id = str(order_id or "").strip()
    if not contract_ids or not order_id:
        return 0
    with get_engine().begin() as conn:
        if not (
            _table_has_column(conn, "finished_goods_data", "合同号")
            and _table_has_column(conn, "finished_goods_data", "占用订单号")
            and _table_has_column(conn, "finished_goods_data", "状态")
        ):
            return 0
        ret = conn.execute(
            text(
                "UPDATE finished_goods_data "
                "SET `占用订单号` = :order_id "
                "WHERE TRIM(COALESCE(`合同号`, '')) COLLATE utf8mb4_general_ci IN :contract_ids "
                "AND TRIM(COALESCE(`状态`, '')) <> '已出库' "
                "AND (COALESCE(TRIM(`占用订单号`), '') = '' OR TRIM(`占用订单号`) = :order_id)"
            ).bindparams(bindparam("contract_ids", expanding=True)),
            {"contract_ids": contract_ids, "order_id": order_id},
        )
    return int(ret.rowcount or 0)


def _table_has_column(conn, table_name: str, column_name: str) -> bool:
    try:
        return conn.execute(
            text(f"SHOW COLUMNS FROM `{table_name}` LIKE :column_name"),
            {"column_name": column_name},
        ).fetchone() is not None
    except Exception:
        return False


def _ensure_plan_import_order_column(conn) -> None:
    if not _table_has_column(conn, "plan_import", "订单号"):
        conn.execute(text("ALTER TABLE plan_import ADD COLUMN `订单号` VARCHAR(100) DEFAULT '' AFTER `合同号`"))


def _table_exists(conn, table_name: str) -> bool:
    try:
        return conn.execute(
            text("SHOW TABLES LIKE :table_name"),
            {"table_name": table_name},
        ).fetchone() is not None
    except Exception:
        return False


def _clear_planning_related_caches() -> None:
    for fn in (get_factory_plan, get_factory_plan_v2, get_orders):
        if hasattr(fn, "cache_clear"):
            fn.cache_clear()
    try:
        import crud.inventory
        if hasattr(crud.inventory.get_data, "cache_clear"):
            crud.inventory.get_data.cache_clear()
    except Exception:
        pass


def _cleanup_cancelled_contract_links(conn, contract_id: str, operator: str = "") -> dict[str, int]:
    contract_id = str(contract_id or "").strip()
    stats = {"factory_plan": 0, "production_queue": 0, "rush_order_queue": 0, "units": 0, "production_history_ledger": 0}
    if not contract_id:
        return stats

    ret = conn.execute(
        text("UPDATE factory_plan SET `状态` = '已取消' WHERE TRIM(COALESCE(`合同号`, '')) = :cid"),
        {"cid": contract_id},
    )
    stats["factory_plan"] = int(ret.rowcount or 0)

    if _table_exists(conn, "production_queue") and _table_has_column(conn, "production_queue", "contract_no"):
        ret = conn.execute(
            text("""
                DELETE FROM production_queue
                WHERE TRIM(COALESCE(contract_no, '')) = :cid
                  AND status = 'Waiting'
            """),
            {"cid": contract_id},
        )
        stats["production_queue"] = int(ret.rowcount or 0)

    if _table_exists(conn, "rush_order_queue") and _table_has_column(conn, "rush_order_queue", "contract_no"):
        set_parts = ["status = 'deleted'"]
        params: dict[str, Any] = {"cid": contract_id}
        if _table_has_column(conn, "rush_order_queue", "updated_by"):
            set_parts.append("updated_by = :operator")
            params["operator"] = str(operator or "")
        ret = conn.execute(
            text(f"""
                UPDATE rush_order_queue
                SET {', '.join(set_parts)}
                WHERE TRIM(COALESCE(contract_no, '')) = :cid
                  AND status = 'pending'
            """),
            params,
        )
        stats["rush_order_queue"] = int(ret.rowcount or 0)

    if _table_exists(conn, "units") and _table_has_column(conn, "units", "contract_no"):
        set_parts = [
            "contract_no = NULL",
            "customer = NULL",
            "dealer_name = NULL",
            "due_date = NULL",
            "sales_id = NULL",
            "order_remark = NULL",
            "is_locked = 0",
        ]
        if _table_has_column(conn, "units", "dealer_id"):
            set_parts.append("dealer_id = NULL")
        if _table_has_column(conn, "units", "is_contract_pinned"):
            set_parts.append("is_contract_pinned = 0")
        if _table_has_column(conn, "units", "locked_by"):
            set_parts.append("locked_by = NULL")
        if _table_has_column(conn, "units", "locked_at"):
            set_parts.append("locked_at = NULL")
        if _table_has_column(conn, "units", "updated_at"):
            set_parts.append("updated_at = NOW()")
        ret = conn.execute(
            text(f"""
                UPDATE units
                SET {', '.join(set_parts)}
                WHERE TRIM(COALESCE(contract_no, '')) = :cid
            """),
            {"cid": contract_id},
        )
        stats["units"] = int(ret.rowcount or 0)

    if (
        _table_exists(conn, "production_history_ledger")
        and _table_has_column(conn, "production_history_ledger", "contract_no")
        and _table_has_column(conn, "production_history_ledger", "status")
    ):
        set_parts = ["status = 'Cancelled'"]
        if _table_has_column(conn, "production_history_ledger", "completed_at"):
            set_parts.append("completed_at = NOW()")
        ret = conn.execute(
            text(f"""
                UPDATE production_history_ledger
                SET {', '.join(set_parts)}
                WHERE status = 'In_Production'
                  AND TRIM(COALESCE(contract_no, '')) = :cid
            """),
            {"cid": contract_id},
        )
        stats["production_history_ledger"] = int(ret.rowcount or 0)

    return stats


def _to_contract_date_text(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    parsed = pd.to_datetime(raw, errors="coerce")
    if pd.isna(parsed):
        return raw
    return parsed.strftime("%Y-%m-%d")


def _to_positive_int(value: object) -> int:
    try:
        qty = int(float(value or 0))
    except Exception:
        qty = 0
    return max(0, qty)


def _normalize_contract_status(value: object) -> str:
    status = str(value or "").strip()
    if status in {"", "未下单"}:
        return "待规划"
    return status


def _safe_json_text(value: object) -> str:
    if isinstance(value, str):
        parsed = parse_alloc_dict(value)
        if parsed:
            return json.dumps(parsed, ensure_ascii=False)
        raw = value.strip()
        return raw if raw.startswith("{") else "{}"
    return json.dumps(parse_alloc_dict(value), ensure_ascii=False)


def _contract_model_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        model = str(row.get("机型") or "").strip()
        if not model:
            continue
        counts[model] = counts.get(model, 0) + _to_positive_int(row.get("排产数量"))
    return counts


def _contract_demand_text(rows: list[dict[str, Any]]) -> tuple[str, int]:
    counts = _contract_model_counts(rows)
    parts = [f"{model}:{qty}" for model, qty in counts.items() if qty > 0]
    return ";".join(parts), sum(qty for qty in counts.values() if qty > 0)


def _build_contract_edit_rows(
    contract_id: str,
    payload: ContractEditPayload,
    existing_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], str]:
    due_date = _to_contract_date_text(payload.要求交期)
    if not due_date:
        raise HTTPException(status_code=422, detail="要求交期不能为空")

    first_status = _normalize_contract_status(existing_rows[0].get("状态") if existing_rows else "待规划")
    status_by_model: dict[str, str] = {}
    source_by_model: dict[str, object] = {}
    order_id = ""
    for row in existing_rows:
        model = str(row.get("机型") or "").strip()
        if model and model not in status_by_model:
            status_by_model[model] = _normalize_contract_status(row.get("状态"))
        if model and model not in source_by_model:
            source_by_model[model] = parse_alloc_dict(row.get("指定批次/来源"))
        if not order_id:
            order_id = str(row.get("订单号") or "").strip()

    merged: dict[str, dict[str, Any]] = {}
    for item in payload.items:
        model = str(item.机型 or "").strip()
        qty = _to_positive_int(item.排产数量)
        if not model:
            continue
        if qty <= 0:
            raise HTTPException(status_code=422, detail=f"机型 {model} 的排产数量必须大于 0")
        note = str(item.备注 or "").strip()
        if model not in merged:
            merged[model] = {
                "合同号": contract_id,
                "机型": model,
                "排产数量": 0,
                "要求交期": due_date,
                "状态": status_by_model.get(model, first_status),
                "备注": "",
                "客户名": str(payload.客户名 or "").strip(),
                "代理商": str(payload.代理商 or "").strip(),
                "指定批次/来源": source_by_model.get(model, {}),
                "订单号": order_id,
            }
        merged[model]["排产数量"] += qty
        if note:
            existing_note = str(merged[model].get("备注") or "").strip()
            if not existing_note:
                merged[model]["备注"] = note
            elif note not in existing_note.split("；"):
                merged[model]["备注"] = f"{existing_note}；{note}"

    rows = list(merged.values())
    if not rows:
        raise HTTPException(status_code=422, detail="机型明细无有效数据")
    return rows, order_id


def _row_to_contract_edit_dict(row: Any) -> dict[str, Any]:
    data = dict(row)
    data["指定批次/来源"] = parse_alloc_dict(data.get("指定批次/来源"))
    return data


def _upsert_factory_plan_contract_rows(conn, contract_id: str, new_rows: list[dict[str, Any]]) -> dict[str, int]:
    old_rows = conn.execute(
        text("""
            SELECT `id`, `机型`
            FROM factory_plan
            WHERE TRIM(COALESCE(`合同号`, '')) = :cid
            ORDER BY `id` ASC
        """),
        {"cid": contract_id},
    ).mappings().all()
    reusable_by_model: dict[str, list[int]] = {}
    all_ids: set[int] = set()
    for row in old_rows:
        row_id = int(row["id"])
        all_ids.add(row_id)
        model = str(row.get("机型") or "").strip()
        reusable_by_model.setdefault(model, []).append(row_id)

    updated = 0
    inserted = 0
    used_ids: set[int] = set()
    for row in new_rows:
        model = str(row.get("机型") or "").strip()
        existing_id = None
        candidates = reusable_by_model.get(model) or []
        while candidates and existing_id is None:
            candidate = candidates.pop(0)
            if candidate not in used_ids:
                existing_id = candidate

        params = {
            "cid": contract_id,
            "model": model,
            "qty": str(_to_positive_int(row.get("排产数量"))),
            "due": str(row.get("要求交期") or ""),
            "status": _normalize_contract_status(row.get("状态")),
            "remark": str(row.get("备注") or ""),
            "customer": str(row.get("客户名") or ""),
            "dealer": str(row.get("代理商") or ""),
            "source": _safe_json_text(row.get("指定批次/来源")),
            "order_id": str(row.get("订单号") or "").strip() or None,
        }
        if existing_id is not None:
            used_ids.add(existing_id)
            conn.execute(
                text("""
                    UPDATE factory_plan
                    SET `机型` = :model,
                        `排产数量` = :qty,
                        `要求交期` = :due,
                        `状态` = :status,
                        `备注` = :remark,
                        `客户名` = :customer,
                        `代理商` = :dealer,
                        `指定批次/来源` = :source,
                        `订单号` = :order_id
                    WHERE `id` = :id AND `合同号` = :cid
                """),
                {**params, "id": existing_id},
            )
            updated += 1
        else:
            conn.execute(
                text("""
                    INSERT INTO factory_plan
                        (`合同号`, `机型`, `排产数量`, `要求交期`, `状态`, `备注`, `客户名`, `代理商`, `指定批次/来源`, `订单号`)
                    VALUES
                        (:cid, :model, :qty, :due, :status, :remark, :customer, :dealer, :source, :order_id)
                """),
                params,
            )
            inserted += 1

    delete_ids = sorted(all_ids - used_ids)
    deleted = 0
    if delete_ids:
        ret = conn.execute(
            text("DELETE FROM factory_plan WHERE `id` IN :ids AND `合同号` = :cid").bindparams(bindparam("ids", expanding=True)),
            {"ids": delete_ids, "cid": contract_id},
        )
        deleted = int(ret.rowcount or 0)
    return {"updated": updated, "inserted": inserted, "deleted": deleted}


def _sync_sales_order_from_contract_edit(conn, order_id: str, new_rows: list[dict[str, Any]]) -> int:
    order_id = str(order_id or "").strip()
    if not order_id or not _table_exists(conn, "sales_orders"):
        return 0

    linked_rows_raw = conn.execute(
        text("""
            SELECT `合同号`, `机型`, `排产数量`, `客户名`, `代理商`, `指定批次/来源`
            FROM factory_plan
            WHERE TRIM(COALESCE(`订单号`, '')) = :order_id
            ORDER BY `id` ASC
        """),
        {"order_id": order_id},
    ).mappings().all()
    linked_rows = [_row_to_contract_edit_dict(row) for row in linked_rows_raw] or new_rows
    demand_text, total_qty = _contract_demand_text(linked_rows)

    customers = [str(row.get("客户名") or "").strip() for row in linked_rows if str(row.get("客户名") or "").strip()]
    dealers = [str(row.get("代理商") or "").strip() for row in linked_rows if str(row.get("代理商") or "").strip()]
    unique_customers = sorted(set(customers))
    unique_dealers = sorted(set(dealers))

    source_map: dict[str, Any] = {}
    for row in linked_rows:
        model = str(row.get("机型") or "").strip()
        alloc = parse_alloc_dict(row.get("指定批次/来源"))
        if model and alloc:
            source_map[model] = alloc
    set_parts = [
        "`需求机型` = :demand_text",
        "`需求数量` = :total_qty",
        "`指定批次/来源` = :source_json",
    ]
    params: dict[str, Any] = {
        "demand_text": demand_text,
        "total_qty": total_qty,
        "source_json": json.dumps(source_map, ensure_ascii=False),
        "order_id": order_id,
    }
    if len(unique_customers) == 1:
        set_parts.append("`客户名` = :customer")
        params["customer"] = unique_customers[0]
    if len(unique_dealers) == 1:
        set_parts.append("`代理商` = :dealer")
        params["dealer"] = unique_dealers[0]
    ret = conn.execute(
        text(f"""
            UPDATE sales_orders
            SET {', '.join(set_parts)}
            WHERE `订单号` = :order_id
        """),
        params,
    )
    return int(ret.rowcount or 0)


def _load_model_family_map(conn, models: list[str]) -> dict[str, str]:
    clean_models = sorted({str(model or "").strip() for model in models if str(model or "").strip()})
    if not clean_models:
        return {}
    rows = conn.execute(
        text("""
            SELECT model_name, COALESCE(model_family, '') AS model_family
            FROM model_dictionary
            WHERE enabled = 1
              AND TRIM(model_name) IN :models
        """).bindparams(bindparam("models", expanding=True)),
        {"models": clean_models},
    ).mappings().all()
    return {
        str(row.get("model_name") or "").strip(): str(row.get("model_family") or "").strip()
        for row in rows
        if str(row.get("model_name") or "").strip()
    }


def _load_contract_unit_rows(conn, contract_id: str, for_update: bool = False) -> list[dict[str, Any]]:
    if not (_table_exists(conn, "units") and _table_exists(conn, "batches")):
        return []
    lock_sql = "FOR UPDATE" if for_update else ""
    rows = conn.execute(
        text(f"""
            SELECT u.unit_id, u.model_type, u.customer, u.dealer_name, u.due_date, u.sales_id,
                   u.order_remark, u.is_locked, u.batch_id, u.slot_index, u.status,
                   COALESCE(b.status, '') AS batch_status,
                   COALESCE(b.model_type, '') AS batch_model_type,
                   COALESCE(md.model_family, '') AS model_family
            FROM units u
            LEFT JOIN batches b ON b.batch_id = u.batch_id
            LEFT JOIN model_dictionary md ON md.model_name = u.model_type COLLATE utf8mb4_general_ci
            WHERE TRIM(COALESCE(u.contract_no, '')) = :contract_id
            ORDER BY CASE COALESCE(b.status, '')
                        WHEN 'Predicted' THEN 0
                        WHEN 'Confirmed' THEN 1
                        WHEN 'In_Production' THEN 2
                        ELSE 3
                     END,
                     u.batch_id ASC, u.slot_index ASC, u.unit_id ASC
            {lock_sql}
        """),
        {"contract_id": contract_id},
    ).mappings().all()
    result: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        if item.get("due_date") is not None:
            item["due_date"] = _to_contract_date_text(item.get("due_date"))
        result.append(item)
    return result


def _contract_edit_preview_token(contract_id: str, new_rows: list[dict[str, Any]], unit_rows: list[dict[str, Any]]) -> str:
    unit_state = [
        {
            "unit_id": str(row.get("unit_id") or ""),
            "model_type": str(row.get("model_type") or ""),
            "batch_status": str(row.get("batch_status") or ""),
            "is_locked": bool(row.get("is_locked")),
        }
        for row in unit_rows
    ]
    payload = {
        "contract_id": contract_id,
        "new": _contract_model_counts(new_rows),
        "units": unit_state,
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _build_contract_edit_preview(
    conn,
    contract_id: str,
    existing_rows: list[dict[str, Any]],
    new_rows: list[dict[str, Any]],
    for_update: bool = False,
) -> dict[str, Any]:
    old_counts = _contract_model_counts(existing_rows)
    new_counts = _contract_model_counts(new_rows)
    old_models = list(old_counts.keys())
    new_models = list(new_counts.keys())
    family_map = _load_model_family_map(conn, old_models + new_models)
    missing_family = sorted({m for m in old_models + new_models if not family_map.get(m)})
    if missing_family:
        raise HTTPException(status_code=422, detail=f"机型缺少启用的机型族配置：{'、'.join(missing_family)}")

    old_family_counts: dict[str, int] = {}
    new_family_counts: dict[str, int] = {}
    for model, qty in old_counts.items():
        old_family_counts[family_map[model]] = old_family_counts.get(family_map[model], 0) + qty
    for model, qty in new_counts.items():
        new_family_counts[family_map[model]] = new_family_counts.get(family_map[model], 0) + qty

    reduced_families = [
        {
            "family": family,
            "old_qty": old_family_counts.get(family, 0),
            "new_qty": new_family_counts.get(family, 0),
        }
        for family in sorted(set(old_family_counts) | set(new_family_counts))
        if new_family_counts.get(family, 0) < old_family_counts.get(family, 0)
    ]
    increased_families = [
        {
            "family": family,
            "old_qty": old_family_counts.get(family, 0),
            "new_qty": new_family_counts.get(family, 0),
        }
        for family in sorted(set(old_family_counts) | set(new_family_counts))
        if new_family_counts.get(family, 0) > old_family_counts.get(family, 0)
    ]
    blocked_families = reduced_families + increased_families if reduced_families and increased_families else []

    unit_rows = _load_contract_unit_rows(conn, contract_id, for_update=for_update)
    old_demand_text, old_total_qty = _contract_demand_text(existing_rows)
    new_demand_text, new_total_qty = _contract_demand_text(new_rows)
    count_changed = old_counts != new_counts
    model_set_changed = set(old_counts.keys()) != set(new_counts.keys())
    requires_mapping = bool(unit_rows) and (count_changed or model_set_changed)

    assigned_units: set[str] = set()
    assignments = []
    releases = []
    supplements = []
    units_by_family: dict[str, list[dict[str, Any]]] = {}
    for row in unit_rows:
        family = str(row.get("model_family") or "").strip()
        units_by_family.setdefault(family, []).append(row)

    status_rank = {"Predicted": 0, "Confirmed": 1, "In_Production": 2}
    for row in new_rows:
        model = str(row.get("机型") or "").strip()
        qty = _to_positive_int(row.get("排产数量"))
        family = family_map.get(model, "")
        candidates = units_by_family.get(family, [])
        exact = [u for u in candidates if str(u.get("model_type") or "").strip() == model and str(u.get("unit_id") or "") not in assigned_units]
        same_family = [u for u in candidates if str(u.get("unit_id") or "") not in assigned_units and u not in exact]
        ordered = exact + same_family
        for unit in ordered[:qty]:
            assigned_units.add(str(unit.get("unit_id") or ""))
            assignments.append({
                "unit_id": unit.get("unit_id"),
                "from_model": unit.get("model_type"),
                "to_model": model,
                "model_family": family,
                "batch_status": unit.get("batch_status"),
                "batch_id": unit.get("batch_id"),
                "slot_index": unit.get("slot_index"),
                "recommended": True,
            })
        missing = max(0, qty - len(ordered))
        if missing:
            supplements.append({
                "model": model,
                "model_family": family,
                "qty": missing,
                "reason": "新增数量或当前绑定卡片不足，建议进入补排",
            })

    for unit in unit_rows:
        unit_id = str(unit.get("unit_id") or "")
        if unit_id and unit_id not in assigned_units:
            releases.append({
                "unit_id": unit.get("unit_id"),
                "model": unit.get("model_type"),
                "model_family": unit.get("model_family"),
                "batch_status": unit.get("batch_status"),
                "batch_id": unit.get("batch_id"),
                "slot_index": unit.get("slot_index"),
                "recommended": True,
            })
    releases.sort(key=lambda x: (status_rank.get(str(x.get("batch_status") or ""), 9), str(x.get("batch_id") or ""), int(x.get("slot_index") or 0)))

    by_status: dict[str, int] = {}
    for unit in unit_rows:
        status = str(unit.get("batch_status") or "Unknown")
        by_status[status] = by_status.get(status, 0) + 1

    blocked = bool(blocked_families)
    return {
        "blocked": blocked,
        "blocked_reason": "跨机型族替换会改变既有绑定卡片的机型族，不能直接修改；请取消、拆分或新建合同处理" if blocked else "",
        "blocked_families": blocked_families,
        "requires_mapping": requires_mapping and not blocked,
        "preview_token": _contract_edit_preview_token(contract_id, new_rows, unit_rows),
        "diff": {
            "old_demand": {"text": old_demand_text, "quantity": old_total_qty, "counts": old_counts, "family_counts": old_family_counts},
            "new_demand": {"text": new_demand_text, "quantity": new_total_qty, "counts": new_counts, "family_counts": new_family_counts},
            "customer_changed": str(existing_rows[0].get("客户名") or "").strip() != str(new_rows[0].get("客户名") or "").strip(),
            "dealer_changed": str(existing_rows[0].get("代理商") or "").strip() != str(new_rows[0].get("代理商") or "").strip(),
            "due_date_changed": _to_contract_date_text(existing_rows[0].get("要求交期")) != _to_contract_date_text(new_rows[0].get("要求交期")),
            "count_changed": count_changed,
            "model_set_changed": model_set_changed,
        },
        "impact": {
            "bound_units": len(unit_rows),
            "by_status": by_status,
        },
        "unit_plan": {
            "assignments": assignments,
            "releases": releases,
            "supplements": supplements,
        },
        "families": family_map,
    }


def _validate_contract_edit_decision(preview: dict[str, Any], decision: dict[str, Any] | None) -> dict[str, Any]:
    if preview.get("blocked"):
        raise HTTPException(status_code=422, detail=preview.get("blocked_reason") or "合同修改被拦截")
    if not preview.get("requires_mapping"):
        return {"assignments": [], "releases": [], "supplements": []}
    if not isinstance(decision, dict):
        raise HTTPException(status_code=409, detail={"message": "该修改涉及数量变化或增删机型，需要先确认卡片调整方案", "preview": preview})
    if str(decision.get("preview_token") or "") != str(preview.get("preview_token") or ""):
        raise HTTPException(status_code=409, detail={"message": "合同影响预检已过期，请重新预检", "preview": preview})
    plan = decision.get("unit_plan") if isinstance(decision.get("unit_plan"), dict) else decision
    assignments = plan.get("assignments") if isinstance(plan.get("assignments"), list) else []
    releases = plan.get("releases") if isinstance(plan.get("releases"), list) else []
    supplements = plan.get("supplements") if isinstance(plan.get("supplements"), list) else []

    expected = preview.get("unit_plan", {})
    all_units: dict[str, dict[str, Any]] = {}
    for item in [*(expected.get("assignments", []) or []), *(expected.get("releases", []) or [])]:
        unit_id = str(item.get("unit_id") or "").strip()
        if unit_id:
            all_units[unit_id] = item

    new_counts = {
        str(model or "").strip(): _to_positive_int(qty)
        for model, qty in (preview.get("diff", {}).get("new_demand", {}).get("counts", {}) or {}).items()
        if str(model or "").strip()
    }
    valid_models = set(new_counts)
    family_map = preview.get("families", {}) if isinstance(preview.get("families"), dict) else {}

    assigned_ids: set[str] = set()
    released_ids: set[str] = set()
    assigned_by_model: dict[str, int] = {}
    normalized_assignments: list[dict[str, Any]] = []
    for item in assignments:
        if not isinstance(item, dict):
            raise HTTPException(status_code=422, detail="确认方案格式错误")
        unit_id = str(item.get("unit_id") or "").strip()
        model = str(item.get("to_model") or item.get("model") or "").strip()
        if not unit_id or not model:
            raise HTTPException(status_code=422, detail="保留卡片必须包含 unit_id 和目标机型")
        if unit_id not in all_units:
            raise HTTPException(status_code=422, detail=f"卡片 {unit_id} 不属于当前合同，不能纳入确认方案")
        if unit_id in assigned_ids:
            raise HTTPException(status_code=422, detail=f"卡片 {unit_id} 重复保留")
        if model not in valid_models:
            raise HTTPException(status_code=422, detail=f"目标机型 {model} 不在新合同明细中")
        unit_family = str(all_units[unit_id].get("model_family") or "").strip()
        model_family = str(family_map.get(model) or "").strip()
        if not unit_family or unit_family != model_family:
            raise HTTPException(status_code=422, detail=f"卡片 {unit_id} 不能跨机型族改为 {model}")
        assigned_ids.add(unit_id)
        assigned_by_model[model] = assigned_by_model.get(model, 0) + 1
        normalized_assignments.append({
            **item,
            "unit_id": unit_id,
            "from_model": item.get("from_model") or all_units[unit_id].get("from_model") or all_units[unit_id].get("model"),
            "to_model": model,
            "model_family": model_family,
        })

    for item in releases:
        if not isinstance(item, dict):
            raise HTTPException(status_code=422, detail="确认方案格式错误")
        unit_id = str(item.get("unit_id") or "").strip()
        if not unit_id:
            raise HTTPException(status_code=422, detail="释放卡片必须包含 unit_id")
        if unit_id not in all_units:
            raise HTTPException(status_code=422, detail=f"卡片 {unit_id} 不属于当前合同，不能释放")
        if unit_id in released_ids:
            raise HTTPException(status_code=422, detail=f"卡片 {unit_id} 重复释放")
        if unit_id in assigned_ids:
            raise HTTPException(status_code=422, detail=f"卡片 {unit_id} 不能同时保留和释放")
        released_ids.add(unit_id)

    missing_decision = sorted(set(all_units) - assigned_ids - released_ids)
    if missing_decision:
        raise HTTPException(status_code=422, detail=f"还有 {len(missing_decision)} 张已绑定卡片未确认保留或释放")

    normalized_supplements: list[dict[str, Any]] = []
    supplement_by_model: dict[str, int] = {}
    for item in supplements:
        if not isinstance(item, dict):
            raise HTTPException(status_code=422, detail="确认方案格式错误")
        model = str(item.get("model") or item.get("to_model") or "").strip()
        qty = _to_positive_int(item.get("qty"))
        if not model or qty <= 0:
            continue
        if model not in valid_models:
            raise HTTPException(status_code=422, detail=f"补排机型 {model} 不在新合同明细中")
        supplement_by_model[model] = supplement_by_model.get(model, 0) + qty
        normalized_supplements.append({
            **item,
            "model": model,
            "model_family": str(family_map.get(model) or "").strip(),
            "qty": qty,
        })

    for model, total_qty in new_counts.items():
        if assigned_by_model.get(model, 0) + supplement_by_model.get(model, 0) != total_qty:
            raise HTTPException(
                status_code=422,
                detail=f"{model} 的确认数量不等于新合同数量：保留 {assigned_by_model.get(model, 0)} + 补排 {supplement_by_model.get(model, 0)}，合同 {total_qty}",
            )
    extra_models = sorted((set(assigned_by_model) | set(supplement_by_model)) - set(new_counts))
    if extra_models:
        raise HTTPException(status_code=422, detail=f"确认方案包含新合同外机型：{'、'.join(extra_models)}")

    return {
        "assignments": normalized_assignments,
        "releases": [item for item in releases if str(item.get("unit_id") or "").strip()],
        "supplements": normalized_supplements,
    }


def _load_contract_edit_context(conn, contract_id: str, payload: ContractEditPayload, for_update: bool = False) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str, dict[str, Any]]:
    lock_sql = "FOR UPDATE" if for_update else ""
    existing_db_rows = conn.execute(
        text(f"""
            SELECT `id`, `合同号`, `机型`, `排产数量`, `要求交期`, `状态`, `备注`,
                   `客户名`, `代理商`, `指定批次/来源`, `订单号`
            FROM factory_plan
            WHERE TRIM(COALESCE(`合同号`, '')) = :cid
            ORDER BY `id` ASC
            {lock_sql}
        """),
        {"cid": contract_id},
    ).mappings().all()
    if not existing_db_rows:
        raise HTTPException(status_code=404, detail="合同不存在")
    existing_rows = [_row_to_contract_edit_dict(row) for row in existing_db_rows]
    new_rows, order_id = _build_contract_edit_rows(contract_id, payload, existing_rows)
    preview = _build_contract_edit_preview(conn, contract_id, existing_rows, new_rows, for_update=for_update)
    return existing_rows, new_rows, order_id, preview


def _sync_units_from_contract_edit(
    conn,
    contract_id: str,
    new_rows: list[dict[str, Any]],
    order_id: str,
    unit_plan: dict[str, Any] | None = None,
) -> tuple[dict[str, int], list[dict[str, Any]]]:
    stats = {"display_updated": 0, "rebound": 0, "released": 0, "supplement_queued": 0}
    conflicts: list[dict[str, Any]] = []
    if not (_table_exists(conn, "units") and _table_exists(conn, "batches")):
        return stats, conflicts

    first = new_rows[0]
    due_date = _to_contract_date_text(first.get("要求交期")) or None
    ret = conn.execute(
        text("""
            UPDATE units
            SET customer = :customer,
                dealer_name = :dealer,
                due_date = :due_date,
                sales_id = CASE WHEN :order_id <> '' THEN :order_id ELSE sales_id END,
                updated_at = NOW()
            WHERE TRIM(COALESCE(contract_no, '')) = :contract_id
        """),
        {
            "customer": str(first.get("客户名") or "") or None,
            "dealer": str(first.get("代理商") or "") or None,
            "due_date": due_date,
            "order_id": str(order_id or "").strip(),
            "contract_id": contract_id,
        },
    )
    stats["display_updated"] = int(ret.rowcount or 0)

    remark_by_model = {str(row.get("机型") or "").strip(): str(row.get("备注") or "") for row in new_rows}
    row_by_model = {str(row.get("机型") or "").strip(): row for row in new_rows}
    plan = unit_plan or {"assignments": [], "releases": [], "supplements": []}
    assignments = plan.get("assignments") if isinstance(plan.get("assignments"), list) else []
    releases = plan.get("releases") if isinstance(plan.get("releases"), list) else []
    supplements = plan.get("supplements") if isinstance(plan.get("supplements"), list) else []

    if not assignments and not releases and not supplements:
        # 只改客户/代理商/交期/备注，或者没有绑定卡片时，按当前机型更新备注即可。
        for model, remark in remark_by_model.items():
            conn.execute(
                text("""
                    UPDATE units
                    SET order_remark = :remark,
                        updated_at = NOW()
                    WHERE TRIM(COALESCE(contract_no, '')) = :contract_id
                      AND TRIM(COALESCE(model_type, '')) = :model
                """),
                {"contract_id": contract_id, "model": model, "remark": remark or None},
            )
            if (
                _table_exists(conn, "production_history_ledger")
                and _table_has_column(conn, "production_history_ledger", "contract_no")
                and _table_has_column(conn, "production_history_ledger", "model_type")
                and _table_has_column(conn, "production_history_ledger", "order_remark")
            ):
                conn.execute(
                    text("""
                        UPDATE production_history_ledger
                        SET order_remark = :remark
                        WHERE status = 'In_Production'
                          AND TRIM(COALESCE(contract_no, '')) = :contract_id
                          AND TRIM(COALESCE(model_type, '')) = :model
                    """),
                    {"contract_id": contract_id, "model": model, "remark": remark or None},
                )
        if _table_exists(conn, "production_history_ledger"):
            ledger_sets = []
            ledger_params: dict[str, Any] = {"contract_id": contract_id}
            if _table_has_column(conn, "production_history_ledger", "customer"):
                ledger_sets.append("customer = :customer")
                ledger_params["customer"] = str(first.get("客户名") or "") or None
            if _table_has_column(conn, "production_history_ledger", "dealer_name"):
                ledger_sets.append("dealer_name = :dealer")
                ledger_params["dealer"] = str(first.get("代理商") or "") or None
            if ledger_sets:
                conn.execute(
                    text(f"""
                        UPDATE production_history_ledger
                        SET {', '.join(ledger_sets)}
                        WHERE status = 'In_Production'
                          AND TRIM(COALESCE(contract_no, '')) = :contract_id
                    """),
                    ledger_params,
                )
        return stats, conflicts

    for item in assignments:
        unit_id = str(item.get("unit_id") or "").strip()
        model = str(item.get("to_model") or item.get("model") or "").strip()
        if not unit_id or not model:
            continue
        source_row = row_by_model.get(model) or first
        conn.execute(
            text("""
                UPDATE units
                SET contract_no = :contract_id,
                    customer = :customer,
                    dealer_name = :dealer,
                    due_date = :due_date,
                    sales_id = :order_id,
                    order_remark = :remark,
                    model_type = :model,
                    updated_at = NOW()
                WHERE unit_id = :unit_id
                  AND TRIM(COALESCE(contract_no, '')) = :contract_id
            """),
            {
                "unit_id": unit_id,
                "contract_id": contract_id,
                "customer": str(source_row.get("客户名") or "") or None,
                "dealer": str(source_row.get("代理商") or "") or None,
                "due_date": _to_contract_date_text(source_row.get("要求交期")) or None,
                "order_id": str(order_id or "").strip() or None,
                "remark": str(source_row.get("备注") or "") or None,
                "model": model,
            },
        )
        if _table_exists(conn, "production_history_ledger"):
            ledger_sets = []
            ledger_params: dict[str, Any] = {"unit_id": unit_id}
            if _table_has_column(conn, "production_history_ledger", "customer"):
                ledger_sets.append("customer = :customer")
                ledger_params["customer"] = str(source_row.get("客户名") or "") or None
            if _table_has_column(conn, "production_history_ledger", "dealer_name"):
                ledger_sets.append("dealer_name = :dealer")
                ledger_params["dealer"] = str(source_row.get("代理商") or "") or None
            if _table_has_column(conn, "production_history_ledger", "order_remark"):
                ledger_sets.append("order_remark = :remark")
                ledger_params["remark"] = str(source_row.get("备注") or "") or None
            if _table_has_column(conn, "production_history_ledger", "model_type"):
                ledger_sets.append("model_type = :model")
                ledger_params["model"] = model
            if ledger_sets:
                conn.execute(
                    text(f"""
                        UPDATE production_history_ledger
                        SET {', '.join(ledger_sets)}
                        WHERE status = 'In_Production'
                          AND unit_id = :unit_id
                    """),
                    ledger_params,
                )
        stats["rebound"] += 1

    for item in releases:
        unit_id = str(item.get("unit_id") or "").strip()
        if not unit_id:
            continue
        conn.execute(
            text("""
                UPDATE units
                SET contract_no = NULL,
                    customer = NULL,
                    dealer_id = NULL,
                    dealer_name = NULL,
                    due_date = NULL,
                    sales_id = NULL,
                    order_remark = NULL,
                    is_contract_pinned = 0,
                    is_locked = 0,
                    locked_by = NULL,
                    locked_at = NULL,
                    updated_at = NOW()
                WHERE unit_id = :unit_id
                  AND TRIM(COALESCE(contract_no, '')) = :contract_id
            """),
            {"unit_id": unit_id, "contract_id": contract_id},
        )
        stats["released"] += 1
        if _table_exists(conn, "production_history_ledger"):
            ledger_sets = []
            if _table_has_column(conn, "production_history_ledger", "contract_no"):
                ledger_sets.append("contract_no = NULL")
            if _table_has_column(conn, "production_history_ledger", "customer"):
                ledger_sets.append("customer = NULL")
            if _table_has_column(conn, "production_history_ledger", "dealer_name"):
                ledger_sets.append("dealer_name = NULL")
            if _table_has_column(conn, "production_history_ledger", "order_remark"):
                ledger_sets.append("order_remark = NULL")
            if ledger_sets:
                conn.execute(
                    text(f"""
                        UPDATE production_history_ledger
                        SET {', '.join(ledger_sets)}
                        WHERE status = 'In_Production'
                          AND unit_id = :unit_id
                    """),
                    {"unit_id": unit_id},
                )

    queue_rows: list[dict[str, Any]] = []
    for item in supplements:
        model = str(item.get("model") or "").strip()
        qty = _to_positive_int(item.get("qty"))
        source_row = row_by_model.get(model)
        if not source_row or qty <= 0:
            continue
        queue_rows.append({
            "机型": model,
            "合同号": contract_id,
            "客户名": str(source_row.get("客户名") or ""),
            "代理商": str(source_row.get("代理商") or ""),
            "要求交期": _to_contract_date_text(source_row.get("要求交期")),
            "排产数量": qty,
        })
    if queue_rows:
        stats["supplement_queued"] = _upsert_contract_edit_supplement_queue(conn, queue_rows)

    return stats, conflicts


def _sync_plan_import_and_inventory_from_contract_edit(
    conn,
    contract_id: str,
    new_rows: list[dict[str, Any]],
    order_id: str,
) -> tuple[dict[str, int], list[dict[str, Any]]]:
    stats = {"plan_import": 0, "finished_goods": 0}
    conflicts: list[dict[str, Any]] = []
    first = new_rows[0]
    new_models = [str(row.get("机型") or "").strip() for row in new_rows if str(row.get("机型") or "").strip()]

    if _table_exists(conn, "plan_import") and _table_has_column(conn, "plan_import", "合同号"):
        sets = []
        params: dict[str, Any] = {"contract_id": contract_id}
        if _table_has_column(conn, "plan_import", "客户"):
            sets.append("`客户` = :customer")
            params["customer"] = str(first.get("客户名") or "")
        if _table_has_column(conn, "plan_import", "代理商"):
            sets.append("`代理商` = :dealer")
            params["dealer"] = str(first.get("代理商") or "")
        if _table_has_column(conn, "plan_import", "订单号"):
            sets.append("`订单号` = :order_id")
            params["order_id"] = str(order_id or "")
        if sets:
            ret = conn.execute(
                text(f"UPDATE plan_import SET {', '.join(sets)} WHERE TRIM(COALESCE(`合同号`, '')) = :contract_id"),
                params,
            )
            stats["plan_import"] += int(ret.rowcount or 0)
        if _table_has_column(conn, "plan_import", "合同备注"):
            for row in new_rows:
                model = str(row.get("机型") or "").strip()
                if model:
                    conn.execute(
                        text("""
                            UPDATE plan_import
                            SET `合同备注` = :remark
                            WHERE TRIM(COALESCE(`合同号`, '')) = :contract_id
                              AND TRIM(COALESCE(`机型`, '')) = :model
                        """),
                        {"contract_id": contract_id, "model": model, "remark": str(row.get("备注") or "")},
                    )
        if new_models and _table_has_column(conn, "plan_import", "机型"):
            stale = conn.execute(
                text("""
                    SELECT `机型`, COUNT(*) AS qty
                    FROM plan_import
                    WHERE TRIM(COALESCE(`合同号`, '')) = :contract_id
                      AND TRIM(COALESCE(`机型`, '')) NOT IN :models
                    GROUP BY `机型`
                """).bindparams(bindparam("models", expanding=True)),
                {"contract_id": contract_id, "models": new_models},
            ).mappings().all()
            for row in stale:
                conflicts.append({
                    "scope": "plan_import",
                    "type": "imported_model_not_in_contract",
                    "model": str(row.get("机型") or ""),
                    "qty": int(row.get("qty") or 0),
                    "message": f"导入批次中仍有旧机型 {row.get('机型') or '-'} {int(row.get('qty') or 0)} 台，需人工确认",
                })

    if _table_exists(conn, "finished_goods_data") and _table_has_column(conn, "finished_goods_data", "合同号"):
        sets = []
        params = {"contract_id": contract_id}
        if _table_has_column(conn, "finished_goods_data", "客户"):
            sets.append("`客户` = :customer")
            params["customer"] = str(first.get("客户名") or "")
        if _table_has_column(conn, "finished_goods_data", "代理商"):
            sets.append("`代理商` = :dealer")
            params["dealer"] = str(first.get("代理商") or "")
        if sets:
            status_guard = "AND TRIM(COALESCE(`状态`, '')) <> '已出库'" if _table_has_column(conn, "finished_goods_data", "状态") else ""
            ret = conn.execute(
                text(f"""
                    UPDATE finished_goods_data
                    SET {', '.join(sets)}
                    WHERE TRIM(COALESCE(`合同号`, '')) = :contract_id
                    {status_guard}
                """),
                params,
            )
            stats["finished_goods"] += int(ret.rowcount or 0)
        if _table_has_column(conn, "finished_goods_data", "合同备注"):
            for row in new_rows:
                model = str(row.get("机型") or "").strip()
                if model:
                    conn.execute(
                        text("""
                            UPDATE finished_goods_data
                            SET `合同备注` = :remark
                            WHERE TRIM(COALESCE(`合同号`, '')) = :contract_id
                              AND TRIM(COALESCE(`机型`, '')) = :model
                              AND TRIM(COALESCE(`状态`, '')) <> '已出库'
                        """),
                        {"contract_id": contract_id, "model": model, "remark": str(row.get("备注") or "")},
                    )
        if new_models and _table_has_column(conn, "finished_goods_data", "机型"):
            stale = conn.execute(
                text("""
                    SELECT `机型`, `状态`, COUNT(*) AS qty
                    FROM finished_goods_data
                    WHERE TRIM(COALESCE(`合同号`, '')) = :contract_id
                      AND TRIM(COALESCE(`机型`, '')) NOT IN :models
                    GROUP BY `机型`, `状态`
                """).bindparams(bindparam("models", expanding=True)),
                {"contract_id": contract_id, "models": new_models},
            ).mappings().all()
            for row in stale:
                conflicts.append({
                    "scope": "finished_goods_data",
                    "type": "physical_model_not_in_contract",
                    "model": str(row.get("机型") or ""),
                    "status": str(row.get("状态") or ""),
                    "qty": int(row.get("qty") or 0),
                    "message": f"库存/实物记录中仍有旧机型 {row.get('机型') or '-'} {int(row.get('qty') or 0)} 台，状态 {row.get('状态') or '-'}，不自动改实物机型",
                })
    return stats, conflicts


def _sync_rush_queue_from_contract_edit(
    conn,
    contract_id: str,
    new_rows: list[dict[str, Any]],
    operator: str,
) -> tuple[dict[str, int], list[dict[str, Any]]]:
    stats = {"pending_rebuilt": 0, "pending_deleted": 0}
    conflicts: list[dict[str, Any]] = []
    if not _table_exists(conn, "rush_order_queue"):
        return stats, conflicts

    total_queue_count = conn.execute(
        text("""
            SELECT COUNT(*) FROM rush_order_queue
            WHERE TRIM(COALESCE(`contract_no`, '')) = :contract_id
        """),
        {"contract_id": contract_id},
    ).scalar() or 0
    if int(total_queue_count) <= 0:
        return stats, conflicts

    pending_count = conn.execute(
        text("""
            SELECT COUNT(*) FROM rush_order_queue
            WHERE TRIM(COALESCE(`contract_no`, '')) = :contract_id
              AND `status` = 'pending'
        """),
        {"contract_id": contract_id},
    ).scalar() or 0
    new_models = [str(row.get("机型") or "").strip() for row in new_rows if str(row.get("机型") or "").strip()]
    processed_by_model: dict[str, int] = {}
    processed_rows = conn.execute(
        text("""
            SELECT model_type, COUNT(*) AS qty
            FROM rush_order_queue
            WHERE TRIM(COALESCE(`contract_no`, '')) = :contract_id
              AND `status` NOT IN ('pending', 'deleted')
            GROUP BY model_type
        """),
        {"contract_id": contract_id},
    ).mappings().all()
    for row in processed_rows:
        model = str(row.get("model_type") or "").strip()
        if model:
            processed_by_model[model] = int(row.get("qty") or 0)

    if new_models:
        stale = conn.execute(
            text("""
                SELECT model_type, status, COUNT(*) AS qty
                FROM rush_order_queue
                WHERE TRIM(COALESCE(`contract_no`, '')) = :contract_id
                  AND `status` NOT IN ('pending', 'deleted')
                  AND TRIM(COALESCE(model_type, '')) NOT IN :models
                GROUP BY model_type, status
            """).bindparams(bindparam("models", expanding=True)),
            {"contract_id": contract_id, "models": new_models},
        ).mappings().all()
        for row in stale:
            conflicts.append({
                "scope": "rush_order_queue",
                "type": "processed_rush_model_not_in_contract",
                "model": str(row.get("model_type") or ""),
                "status": str(row.get("status") or ""),
                "qty": int(row.get("qty") or 0),
                "message": f"已处理急单中仍有旧机型 {row.get('model_type') or '-'}，状态 {row.get('status') or '-'}，需人工处理",
            })

    if int(pending_count) > 0:
        ret = conn.execute(
            text("""
                UPDATE rush_order_queue
                SET `status` = 'deleted',
                    `updated_by` = :operator
                WHERE TRIM(COALESCE(`contract_no`, '')) = :contract_id
                  AND `status` = 'pending'
            """),
            {"contract_id": contract_id, "operator": str(operator or "")},
        )
        stats["pending_deleted"] = int(ret.rowcount or 0)

    insert_rows = []
    for row in new_rows:
        model = str(row.get("机型") or "").strip()
        qty = max(0, _to_positive_int(row.get("排产数量")) - processed_by_model.get(model, 0))
        for _ in range(qty):
            insert_rows.append({
                "contract_no": contract_id,
                "customer": str(row.get("客户名") or ""),
                "dealer_name": str(row.get("代理商") or ""),
                "model_type": model,
                "due_date": _to_contract_date_text(row.get("要求交期")) or None,
                "remark": str(row.get("备注") or ""),
                "source": "contract-edit",
                "status": "pending",
                "created_by": str(operator or ""),
                "updated_by": str(operator or ""),
            })
    if insert_rows:
        conn.execute(
            text("""
                INSERT INTO rush_order_queue
                    (contract_no, customer, dealer_name, model_type, due_date, remark, source, status, created_by, updated_by)
                VALUES
                    (:contract_no, :customer, :dealer_name, :model_type, :due_date, :remark, :source, :status, :created_by, :updated_by)
            """),
            insert_rows,
        )
        stats["pending_rebuilt"] = len(insert_rows)
    return stats, conflicts


def _sync_production_queue_from_contract_edit(conn, contract_id: str, new_rows: list[dict[str, Any]]) -> dict[str, int]:
    stats = {"waiting_updated": 0, "waiting_deleted": 0}
    if not (_table_exists(conn, "production_queue") and _table_has_column(conn, "production_queue", "contract_no")):
        return stats
    new_models = [str(row.get("机型") or "").strip() for row in new_rows if str(row.get("机型") or "").strip()]
    first = new_rows[0]

    set_parts = []
    params: dict[str, Any] = {"contract_id": contract_id}
    if _table_has_column(conn, "production_queue", "customer"):
        set_parts.append("customer = :customer")
        params["customer"] = str(first.get("客户名") or "")
    if _table_has_column(conn, "production_queue", "dealer"):
        set_parts.append("dealer = :dealer")
        params["dealer"] = str(first.get("代理商") or "")
    if _table_has_column(conn, "production_queue", "dealer_name"):
        set_parts.append("dealer_name = :dealer")
        params["dealer"] = str(first.get("代理商") or "")
    if _table_has_column(conn, "production_queue", "due_date"):
        set_parts.append("due_date = :due_date")
        params["due_date"] = _to_contract_date_text(first.get("要求交期")) or None
    if _table_has_column(conn, "production_queue", "payload"):
        set_parts.append(
            "payload = JSON_SET(COALESCE(payload, JSON_OBJECT()), '$.customer', :payload_customer, '$.dealer_name', :payload_dealer, '$.due_date', :payload_due_date)"
        )
        params["payload_customer"] = str(first.get("客户名") or "")
        params["payload_dealer"] = str(first.get("代理商") or "")
        params["payload_due_date"] = _to_contract_date_text(first.get("要求交期"))
    if set_parts:
        ret = conn.execute(
            text(f"""
                UPDATE production_queue
                SET {', '.join(set_parts)}
                WHERE TRIM(COALESCE(contract_no, '')) = :contract_id
                  AND status = 'Waiting'
            """),
            params,
        )
        stats["waiting_updated"] = int(ret.rowcount or 0)

    if new_models:
        ret = conn.execute(
            text("""
                DELETE FROM production_queue
                WHERE TRIM(COALESCE(contract_no, '')) = :contract_id
                  AND status = 'Waiting'
                  AND TRIM(COALESCE(model_type, '')) NOT IN :models
            """).bindparams(bindparam("models", expanding=True)),
            {"contract_id": contract_id, "models": new_models},
        )
        stats["waiting_deleted"] = int(ret.rowcount or 0)

    if _table_has_column(conn, "production_queue", "quantity_remaining"):
        counts = _contract_model_counts(new_rows)
        for model, qty in counts.items():
            conn.execute(
                text("""
                    UPDATE production_queue
                    SET quantity_remaining = LEAST(quantity_remaining, :qty)
                    WHERE TRIM(COALESCE(contract_no, '')) = :contract_id
                      AND status = 'Waiting'
                      AND TRIM(COALESCE(model_type, '')) = :model
                """),
                {"contract_id": contract_id, "model": model, "qty": qty},
            )
    return stats


def _upsert_contract_edit_supplement_queue(conn, rows: list[dict[str, Any]]) -> int:
    if not rows or not (_table_exists(conn, "production_queue") and _table_has_column(conn, "production_queue", "contract_no")):
        return 0
    inserted_or_updated = 0
    has_qty = _table_has_column(conn, "production_queue", "quantity_remaining")
    has_customer = _table_has_column(conn, "production_queue", "customer")
    has_dealer = _table_has_column(conn, "production_queue", "dealer")
    has_dealer_name = _table_has_column(conn, "production_queue", "dealer_name")
    has_due_date = _table_has_column(conn, "production_queue", "due_date")
    has_payload = _table_has_column(conn, "production_queue", "payload")
    has_priority = _table_has_column(conn, "production_queue", "priority")

    for row in rows:
        qty = _to_positive_int(row.get("排产数量"))
        model = str(row.get("机型") or "").strip()
        contract_no = str(row.get("合同号") or "").strip()
        if qty <= 0 or not model or not contract_no:
            continue

        params = {
            "qty": qty,
            "model": model,
            "contract_no": contract_no,
            "customer": str(row.get("客户名") or ""),
            "dealer": str(row.get("代理商") or ""),
            "due_date": _to_contract_date_text(row.get("要求交期")) or None,
            "payload": json.dumps({
                "customer": str(row.get("客户名") or ""),
                "dealer_name": str(row.get("代理商") or ""),
                "due_date": _to_contract_date_text(row.get("要求交期")) or None,
                "quantity_remaining": qty,
                "source": "contract-edit",
            }, ensure_ascii=False),
        }

        if has_qty:
            set_parts = ["quantity_remaining = :qty"]
            if has_customer:
                set_parts.append("customer = :customer")
            if has_dealer:
                set_parts.append("dealer = :dealer")
            if has_dealer_name:
                set_parts.append("dealer_name = :dealer")
            if has_due_date:
                set_parts.append("due_date = :due_date")
            if has_payload:
                set_parts.append("payload = :payload")
            ret = conn.execute(
                text(f"""
                    UPDATE production_queue
                    SET {', '.join(set_parts)}
                    WHERE TRIM(COALESCE(contract_no, '')) = :contract_no
                      AND TRIM(COALESCE(model_type, '')) = :model
                      AND status = 'Waiting'
                """),
                params,
            )
            if int(ret.rowcount or 0) == 0:
                cols = ["model_type", "contract_no", "quantity_remaining", "status"]
                vals = [":model", ":contract_no", ":qty", "'Waiting'"]
                if has_customer:
                    cols.append("customer")
                    vals.append(":customer")
                if has_dealer:
                    cols.append("dealer")
                    vals.append(":dealer")
                if has_dealer_name:
                    cols.append("dealer_name")
                    vals.append(":dealer")
                if has_due_date:
                    cols.append("due_date")
                    vals.append(":due_date")
                if has_payload:
                    cols.append("payload")
                    vals.append(":payload")
                if has_priority:
                    cols.append("priority")
                    vals.append("0")
                conn.execute(
                    text(f"INSERT INTO production_queue ({', '.join(cols)}) VALUES ({', '.join(vals)})"),
                    params,
                )
            inserted_or_updated += qty
        else:
            cols = ["model_type", "contract_no", "status"]
            vals = [":model", ":contract_no", "'Waiting'"]
            if has_payload:
                cols.append("payload")
                vals.append(":payload")
            if has_priority:
                cols.append("priority")
                vals.append("0")
            conn.execute(
                text(f"INSERT INTO production_queue ({', '.join(cols)}) VALUES ({', '.join(vals)})"),
                params,
            )
            inserted_or_updated += qty
    return inserted_or_updated


def _sync_order_to_units_and_import(contract_ids: list[str], order_id: str) -> dict[str, int]:
    contract_ids = _clean_contract_ids(contract_ids)
    order_id = str(order_id or "").strip()
    if not contract_ids or not order_id:
        return {"units": 0, "plan_import": 0}

    with get_engine().begin() as conn:
        conflicts = conn.execute(
            text(
                "SELECT DISTINCT contract_no, sales_id FROM units "
                "WHERE TRIM(COALESCE(contract_no, '')) COLLATE utf8mb4_general_ci IN :contract_ids "
                "AND COALESCE(TRIM(sales_id), '') <> '' "
                "AND TRIM(sales_id) <> :order_id"
            ).bindparams(bindparam("contract_ids", expanding=True)),
            {"contract_ids": contract_ids, "order_id": order_id},
        ).fetchall()
        if conflicts:
            conflict_contract = str(conflicts[0][0] or "").strip()
            conflict_order = str(conflicts[0][1] or "").strip()
            raise HTTPException(status_code=422, detail=f"合同 {conflict_contract} 已绑定其他订单 {conflict_order}")

        unit_ret = conn.execute(
            text(
                "UPDATE units SET sales_id = :order_id, updated_at = NOW() "
                "WHERE TRIM(COALESCE(contract_no, '')) COLLATE utf8mb4_general_ci IN :contract_ids "
                "AND (sales_id IS NULL OR TRIM(sales_id) = '' OR TRIM(sales_id) = :order_id)"
            ).bindparams(bindparam("contract_ids", expanding=True)),
            {"contract_ids": contract_ids, "order_id": order_id},
        )

        finished_goods_rows = 0
        if _table_has_column(conn, "finished_goods_data", "流水号") and _table_has_column(conn, "finished_goods_data", "占用订单号"):
            unit_serial_rows = conn.execute(
                text(
                    "SELECT serial_no, forecast_serial_no, unit_id FROM units "
                    "WHERE TRIM(COALESCE(contract_no, '')) COLLATE utf8mb4_general_ci IN :contract_ids"
                ).bindparams(bindparam("contract_ids", expanding=True)),
                {"contract_ids": contract_ids},
            ).fetchall()
            serials: list[str] = []
            seen_serials: set[str] = set()
            for row in unit_serial_rows:
                for value in row:
                    sn = str(value or "").strip()
                    if not sn or sn in seen_serials:
                        continue
                    seen_serials.add(sn)
                    serials.append(sn)
            if serials:
                set_sql = "`占用订单号` = :order_id"
                if _table_has_column(conn, "finished_goods_data", "订单号"):
                    set_sql += ", `订单号` = :order_id"
                fg_ret = conn.execute(
                    text(
                        f"UPDATE finished_goods_data SET {set_sql} "
                        "WHERE TRIM(COALESCE(`流水号`, '')) IN :serials "
                        "AND TRIM(COALESCE(`状态`, '')) <> '已出库' "
                        "AND (COALESCE(TRIM(`占用订单号`), '') = '' OR TRIM(`占用订单号`) = :order_id)"
                    ).bindparams(bindparam("serials", expanding=True)),
                    {"serials": serials, "order_id": order_id},
                )
                finished_goods_rows = int(fg_ret.rowcount or 0)

        plan_import_rows = 0
        if _table_has_column(conn, "plan_import", "合同号"):
            _ensure_plan_import_order_column(conn)
            import_conflicts = conn.execute(
                text(
                    "SELECT DISTINCT `合同号`, `订单号` FROM plan_import "
                    "WHERE TRIM(COALESCE(`合同号`, '')) COLLATE utf8mb4_general_ci IN :contract_ids "
                    "AND COALESCE(TRIM(`订单号`), '') <> '' "
                    "AND TRIM(`订单号`) <> :order_id"
                ).bindparams(bindparam("contract_ids", expanding=True)),
                {"contract_ids": contract_ids, "order_id": order_id},
            ).fetchall()
            if import_conflicts:
                conflict_contract = str(import_conflicts[0][0] or "").strip()
                conflict_order = str(import_conflicts[0][1] or "").strip()
                raise HTTPException(status_code=422, detail=f"合同 {conflict_contract} 已绑定其他订单 {conflict_order}")

            import_ret = conn.execute(
                text(
                    "UPDATE plan_import SET `订单号` = :order_id "
                    "WHERE TRIM(COALESCE(`合同号`, '')) COLLATE utf8mb4_general_ci IN :contract_ids "
                    "AND (COALESCE(TRIM(`订单号`), '') = '' OR TRIM(`订单号`) = :order_id)"
                ).bindparams(bindparam("contract_ids", expanding=True)),
                {"contract_ids": contract_ids, "order_id": order_id},
            )
            plan_import_rows = int(import_ret.rowcount or 0)

    return {"units": int(unit_ret.rowcount or 0), "plan_import": plan_import_rows, "finished_goods_data": finished_goods_rows}


def _link_contracts_to_order(contract_ids: list[str], order_id: str, status: str | None = "已转订单") -> int:
    if not contract_ids:
        return 0
    
    with get_engine().begin() as conn:
        # 1. 检查是否存在冲突
        existing = conn.execute(
            text(
                "SELECT `合同号`, `订单号` FROM factory_plan "
                "WHERE TRIM(COALESCE(`合同号`, '')) COLLATE utf8mb4_general_ci IN :cids "
                "AND COALESCE(TRIM(`订单号`), '') <> '' AND TRIM(`订单号`) <> :oid"
            ).bindparams(bindparam("cids", expanding=True)),
            {"cids": contract_ids, "oid": order_id}
        ).fetchall()
        
        if existing:
            conflict_contract = str(existing[0][0] or "").strip()
            conflict_order = str(existing[0][1] or "").strip()
            raise HTTPException(status_code=422, detail=f"合同 {conflict_contract} 已绑定其他订单 {conflict_order}")
            
        # 2. 直接使用 SQL UPDATE 更新状态和订单号，避免全表覆盖导致其他合同丢失
        set_status = ""
        params = {"oid": order_id, "cids": contract_ids}
        if status:
            set_status = ", `状态` = :status"
            params["status"] = status
            
        ret = conn.execute(
            text(f"UPDATE factory_plan SET `订单号` = :oid {set_status} "
                 "WHERE TRIM(COALESCE(`合同号`, '')) COLLATE utf8mb4_general_ci IN :cids").bindparams(bindparam("cids", expanding=True)),
            params
        )
        updated_count = int(ret.rowcount or 0)

    # 3. 清理缓存
    import crud.planning
    if hasattr(crud.planning.get_factory_plan, "cache_clear"):
        crud.planning.get_factory_plan.cache_clear()
    if hasattr(crud.planning.get_factory_plan_v2, "cache_clear"):
        crud.planning.get_factory_plan_v2.cache_clear()
    
    _sync_order_to_units_and_import(contract_ids, str(order_id))
    _occupy_inventory_for_order(contract_ids, str(order_id))
    return updated_count


def _validate_contracts_available(contract_ids: list[str], order_id: str) -> None:
    if not contract_ids:
        return
    df_plan = get_factory_plan()
    if df_plan.empty:
        raise HTTPException(status_code=422, detail="未找到可绑定的合同")
    mask = df_plan["合同号"].astype(str).isin(contract_ids)
    if not mask.any():
        raise HTTPException(status_code=422, detail="未找到可绑定的合同")
    existing = df_plan.loc[mask, "订单号"].fillna("").astype(str).str.strip()
    conflict = existing[(existing != "") & (existing != str(order_id))]
    if not conflict.empty:
        raise HTTPException(status_code=422, detail="所选合同中存在已绑定其他订单的记录")


def _get_order_contract_machine_rows(order_id: str) -> pd.DataFrame:
    order_id = str(order_id).strip()
    plan_df = get_factory_plan()
    contract_rows = plan_df[plan_df["订单号"].astype(str).str.strip() == order_id].copy() if not plan_df.empty else pd.DataFrame()
    contract_ids = sorted({
        str(x).strip()
        for x in contract_rows.get("合同号", pd.Series(dtype=str)).tolist()
        if str(x).strip()
    })

    inv_df = get_data()
    if inv_df.empty:
        inv_df = pd.DataFrame(columns=["流水号", "合同号", "占用订单号", "状态", "机型"])
    for col in ["流水号", "合同号", "占用订单号", "状态", "机型", "批次号", "客户", "代理商", "合同备注"]:
        if col not in inv_df.columns:
            inv_df[col] = ""

    linked_rows = pd.DataFrame(columns=inv_df.columns)
    if contract_ids:
        linked_rows = inv_df[inv_df["合同号"].astype(str).str.strip().isin(contract_ids)].copy()

    # 补充：按订单所需机型，额外拉取库存中尚未被任何订单占用的空闲机台。
    # 不限于合同号匹配，没有合同号的空闲机台也可以被配货。
    needed_models: list[tuple[str, bool]] = []
    if not contract_rows.empty:
        seen_model_keys: set[tuple[str, bool]] = set()
        for _, row in contract_rows.iterrows():
            model = str(row.get("机型", "") or "").strip()
            if not model:
                continue
            high = _is_high_model_hint(model, row.get("备注", ""))
            key = (_normalize_alloc_model(model), high)
            if key not in seen_model_keys:
                seen_model_keys.add(key)
                needed_models.append((model, high))
    else:
        # 兜底：factory_plan 中查不到关联合同时，从 sales_orders 的需求文本提取机型
        orders_df = get_orders()
        order_match = orders_df[orders_df["订单号"].astype(str).str.strip() == order_id]
        if not order_match.empty:
            demand_text = str(order_match.iloc[0].get("需求机型", "") or "")
            seen_model_keys: set[tuple[str, bool]] = set()
            for model in _extract_models_from_demand_text(demand_text):
                if not model:
                    continue
                key = (_normalize_alloc_model(model), False)
                if key not in seen_model_keys:
                    seen_model_keys.add(key)
                    needed_models.append((model, False))
    if needed_models:
        free_rows = inv_df[
            (inv_df["占用订单号"].astype(str).str.strip() == "")
            & (inv_df["状态"].astype(str).str.strip() != "已出库")
        ].copy()
        model_mask = pd.Series(False, index=free_rows.index)
        for model, high in needed_models:
            model_mask = model_mask | free_rows.apply(
                lambda row: _machine_matches_model_requirement(row, model, high),
                axis=1,
            )
        model_rows = free_rows[model_mask].copy()
        linked_rows = pd.concat([linked_rows, model_rows], ignore_index=True).drop_duplicates(subset=["流水号"], keep="first")

    occupied_rows = inv_df[inv_df["占用订单号"].astype(str).str.strip() == order_id].copy()
    rows = pd.concat([linked_rows, occupied_rows], ignore_index=True).drop_duplicates(subset=["流水号"], keep="first")

    required_lookup: dict[tuple[str, str], tuple[str, bool]] = {}
    expected_counts: dict[tuple[str, str], int] = {}
    expected_notes: dict[tuple[str, str], str] = {}
    if not contract_rows.empty:
        for _, row in contract_rows.iterrows():
            cid = str(row.get("合同号", "") or "").strip()
            model = str(row.get("机型", "") or "").strip()
            if not cid or not model:
                continue
            note = str(row.get("备注", "") or "").strip()
            key = (cid, model)
            required_lookup[key] = (model, _is_high_model_hint(model, note))
            try:
                qty = int(float(row.get("排产数量", 0) or 0))
            except Exception:
                qty = 0
            expected_counts[key] = expected_counts.get(key, 0) + max(0, qty)
            if note:
                expected_notes[key] = note
    else:
        # 兜底：factory_plan 没有关联合同时，从 sales_orders 取需求数量
        orders_df = get_orders()
        order_match = orders_df[orders_df["订单号"].astype(str).str.strip() == order_id]
        if not order_match.empty:
            demand_text = str(order_match.iloc[0].get("需求机型", "") or "")
            total_qty = int(order_match.iloc[0].get("需求数量", 0) or 0)
            models = _extract_models_from_demand_text(demand_text)
            qty_per_model = max(1, total_qty // len(models)) if models else 0
            for model in models:
                key = (order_id, model)
                required_lookup[key] = (model, False)
                expected_counts[key] = qty_per_model

    placeholders = []
    for (cid, model), qty in expected_counts.items():
        required_model, required_high = required_lookup.get((cid, model), (model, False))
        matched = rows[
            (
                (rows["合同号"].astype(str).str.strip() == cid)
                | (rows["占用订单号"].astype(str).str.strip() == order_id)
                | (rows["合同号"].astype(str).str.strip() == "")
            )
            & rows.apply(
                lambda row: _machine_matches_model_requirement(row, required_model, required_high),
                axis=1,
            )
        ]
        for i in range(max(0, qty - len(matched))):
            placeholders.append({
                "流水号": "",
                "批次号": "",
                "机型": model,
                "状态": "未入库",
                "预计入库时间": "",
                "更新时间": "",
                "占用订单号": "",
                "客户": "",
                "代理商": "",
                "合同备注": expected_notes.get((cid, model), ""),
                "Location_Code": "",
                "合同号": cid,
                "_placeholder": f"{cid}-{model}-{i + 1}",
            })
    if placeholders:
        rows = pd.concat([rows, pd.DataFrame(placeholders)], ignore_index=True)
    return rows

@router.get("/")
def get_planning_data(
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=2000),
    status: str = Query(""),
    contract_id: str = Query(""),
):
    try:
        where_clauses = []
        params: Dict[str, Any] = {"skip": skip, "limit": limit}
        if str(status).strip():
            where_clauses.append("`状态` = :status")
            params["status"] = str(status).strip()
        if str(contract_id).strip():
            where_clauses.append("`合同号` = :contract_id")
            params["contract_id"] = str(contract_id).strip()
        where_sql = f" WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        count_sql = f"SELECT COUNT(*) AS total FROM factory_plan{where_sql}"
        data_sql = (
            "SELECT `id` AS `_idx`, `合同号`, `机型`, `排产数量`, `要求交期`, `状态`, `备注`, `客户名`, `代理商`, `指定批次/来源`, `订单号` "
            f"FROM factory_plan{where_sql} "
            "ORDER BY `id` DESC LIMIT :limit OFFSET :skip"
        )
        with get_engine().connect() as conn:
            total_df = pd.read_sql(text(count_sql), conn, params=params)
            total = int(total_df.iloc[0]["total"]) if not total_df.empty else 0
            df_plan = pd.read_sql(text(data_sql), conn, params=params)

        if "指定批次/来源" in df_plan.columns:
            df_plan["指定批次/来源"] = df_plan["指定批次/来源"].apply(parse_alloc_dict)
        df_plan = df_plan.where(df_plan.notnull(), None)
        return {"data": df_plan.to_dict(orient="records"), "total": total, "skip": skip, "limit": limit}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/orders")
def get_sales_orders(
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=2000),
    status: str = Query(""),
    keyword: str = Query(""),
):
    try:
        where_clauses = []
        params: Dict[str, Any] = {"skip": skip, "limit": limit}
        if str(status).strip():
            where_clauses.append("`status` = :status")
            params["status"] = str(status).strip()
        if str(keyword).strip():
            where_clauses.append("(`订单号` LIKE :kw OR `客户名` LIKE :kw OR `代理商` LIKE :kw)")
            params["kw"] = f"%{str(keyword).strip()}%"
        where_sql = f" WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        count_sql = f"SELECT COUNT(*) AS total FROM sales_orders{where_sql}"
        data_sql = (
            "SELECT `订单号`, `客户名`, `代理商`, `需求机型`, `需求数量`, "
            "DATE_FORMAT(`下单时间`, '%Y-%m-%d') AS `下单时间`, "
            "`备注`, `包装选项`, "
            "DATE_FORMAT(`发货时间`, '%Y-%m-%d') AS `发货时间`, "
            "`指定批次/来源`, `status`, `delete_reason` "
            "FROM sales_orders"
            f"{where_sql} "
            "ORDER BY sales_orders.`下单时间` DESC LIMIT :limit OFFSET :skip"
        )
        with get_engine().connect() as conn:
            total_df = pd.read_sql(text(count_sql), conn, params=params)
            total = int(total_df.iloc[0]["total"]) if not total_df.empty else 0
            df_orders = pd.read_sql(text(data_sql), conn, params=params)

        df_orders = _reconcile_completed_orders(df_orders)
        df_orders = df_orders.where(df_orders.notnull(), None)
        return {"data": df_orders.to_dict(orient="records"), "total": total, "skip": skip, "limit": limit}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/orders")
def create_sales_order_api(
    payload: SalesOrderCreatePayload,
    request: Request,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        _assert_models_in_dictionary(_extract_models_from_demand_text(str(payload.需求机型 or "")))
        df_orders = get_orders()
        order_id = f"SO-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:4].upper()}"
        contract_ids = _clean_contract_ids(payload.contract_ids)
        _validate_contracts_available(contract_ids, order_id)
        new_row = {
            "订单号": order_id,
            "客户名": str(payload.客户名 or ""),
            "代理商": str(payload.代理商 or ""),
            "需求机型": str(payload.需求机型 or ""),
            "需求数量": int(payload.需求数量),
            "下单时间": datetime.now(),
            "备注": str(payload.备注 or ""),
            "包装选项": str(payload.包装选项 or ""),
            "发货时间": pd.to_datetime(payload.发货时间, errors="coerce") if payload.发货时间 else None,
            "指定批次/来源": {},
            "status": "active",
            "delete_reason": "",
        }
        df_orders = pd.concat([df_orders, pd.DataFrame([new_row])], ignore_index=True)
        save_orders(df_orders)
        linked_count = _link_contracts_to_order(contract_ids, order_id) if contract_ids else 0
        append_audit_log(
            module="销售下单",
            action_type="新增",
            biz_type="订单",
            content=(
                f"创建订单：{order_id}；客户：{payload.客户名}；"
                f"需求机型：{str(payload.需求机型 or '').strip() or '未填写'}；"
                f"需求数量：{int(payload.需求数量)}"
                + (f"；绑定合同：{', '.join(contract_ids)}" if contract_ids else "")
            ),
            user_id=current_user.get("username"),
            username=current_operator,
        )
        return {"message": "订单创建成功", "order_id": order_id, "linked_contract_count": linked_count}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"创建订单失败: {e}")


def _clear_sandbox_units_by_order(order_id: str) -> None:
    """
    【缺口2补全】订单删除时，清空沙盘 units 表中该订单关联卡片的合同字段。
    通过 sales_id = order_id 匹配（_sync_order_to_units_and_import 在创建订单时写入）。
    只清空 Predicted 状态的批次中的卡片，已下达/生产中的卡片不动。
    失败不抛异常（fire-and-forget），避免阻塞主流程。
    """
    try:
        with get_engine().begin() as conn:
            conn.execute(
                text("""
                    UPDATE units u
                    JOIN batches b ON b.batch_id = u.batch_id
                    SET u.contract_no = NULL,
                        u.customer    = NULL,
                        u.dealer_name = NULL,
                        u.dealer_id   = NULL,
                        u.due_date    = NULL,
                        u.sales_id    = NULL,
                        u.order_remark = NULL,
                        u.is_locked   = 0,
                        u.locked_by   = NULL,
                        u.locked_at   = NULL,
                        u.updated_at  = NOW()
                    WHERE u.sales_id = :order_id
                      AND b.status = 'Predicted'
                """),
                {"order_id": order_id}
            )
    except Exception as e:
        print(f"Warning: _clear_sandbox_units_by_order failed for order {order_id}: {e}")


def _sync_contract_fields_to_units(
    contract_id: str,
    customer: str = "",
    dealer_name: str = "",
    due_date: str = "",
    model_type: str = "",
    order_remark: str = "",
) -> None:
    """
    【缺口3补全】合同编辑后局部更新沙盘 units 表中的卡片字段。
    通过 contract_no = contract_id 匹配，只更新 Predicted 批次中的卡片。
    失败不抛异常（fire-and-forget），避免阻塞主流程。
    """
    try:
        due_date_val = None
        if due_date:
            try:
                due_date_val = pd.to_datetime(due_date, errors="coerce")
                if pd.isna(due_date_val):
                    due_date_val = None
                else:
                    due_date_val = due_date_val.strftime("%Y-%m-%d")
            except Exception:
                due_date_val = None

        with get_engine().begin() as conn:
            conn.execute(
                text("""
                    UPDATE units u
                    JOIN batches b ON b.batch_id = u.batch_id
                    SET u.customer     = :customer,
                        u.dealer_name  = :dealer_name,
                        u.due_date     = :due_date,
                        u.order_remark = CASE
                            WHEN :order_remark != '' AND (:model_type = '' OR TRIM(COALESCE(u.model_type, '')) = :model_type) THEN :order_remark
                            ELSE u.order_remark
                        END,
                        u.updated_at   = NOW()
                    WHERE u.contract_no = :contract_id
                      AND b.status IN ('Predicted', 'Confirmed', 'In_Production')
                """),
                {
                    "customer": customer or None,
                    "dealer_name": dealer_name or None,
                    "due_date": due_date_val,
                    "model_type": model_type,
                    "order_remark": order_remark,
                    "contract_id": contract_id,
                }
            )
    except Exception as e:
        print(f"Warning: _sync_contract_fields_to_units failed for contract {contract_id}: {e}")



class UnitSyncPayload(BaseModel):
    contract_no: str
    old_model: str
    new_model: str
    order_remark: str
    customer: str = ""
    dealer_name: str = ""


class ContractCancelSyncPayload(BaseModel):
    contract_no: str
    operator: str = "system"
    source: str = "go"


# 内部专用路由，不带常规用户鉴权，内部校验 GO_INTERNAL_TOKEN
internal_router = APIRouter()


def _assert_internal_token(request: Request) -> None:
    config_token = (GO_INTERNAL_TOKEN or "").strip()
    provided_token = (request.headers.get("X-Internal-Token") or "").strip()
    if config_token and provided_token != config_token:
        raise HTTPException(status_code=403, detail="Unauthorized internal request")


@internal_router.patch("/unit-sync")
def internal_sync_unit_api(payload: UnitSyncPayload, request: Request):
    """
    看板反向同步只允许备注回写。合同机型、客户、代理商必须从合同管理修改。
    """
    _assert_internal_token(request)

    try:
        with get_engine().begin() as conn:
            model_to_match = str(payload.old_model or "").strip() or str(payload.new_model or "").strip()
            conn.execute(
                text("""
                    UPDATE factory_plan 
                    SET `备注` = :remark
                    WHERE `合同号` = :contract_no
                      AND `机型` = :model_to_match
                """),
                {
                    "remark": payload.order_remark,
                    "contract_no": payload.contract_no,
                    "model_to_match": model_to_match,
                }
            )
            print(f"[unit-sync] remark-only contract={payload.contract_no} model={model_to_match}")
        
        # 清理缓存，确保主系统刷新后能看到最新数据
        if hasattr(get_factory_plan, "cache_clear"):
            get_factory_plan.cache_clear()
        if hasattr(get_factory_plan_v2, "cache_clear"):
            get_factory_plan_v2.cache_clear()
            
        return {"status": "success"}
    except Exception as e:
        print(f"Internal Sync Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@internal_router.post("/contract-cancel-sync")
def internal_contract_cancel_sync_api(payload: ContractCancelSyncPayload, request: Request):
    """
    Go 侧标记现货导致整单取消后，回到合同域做最终清理、审计和沙盘重算。
    """
    _assert_internal_token(request)
    contract_no = str(payload.contract_no or "").strip()
    if not contract_no:
        raise HTTPException(status_code=422, detail="contract_no 不能为空")
    operator = str(payload.operator or "system").strip() or "system"
    try:
        with get_engine().begin() as conn:
            stats = _cleanup_cancelled_contract_links(conn, contract_no, operator)
        _clear_planning_related_caches()
        recompute_synced = _trigger_sandbox_recompute_sync({"username": operator, "role": "Admin"})
        append_audit_log(
            module="合同管理",
            action_type="取消",
            biz_type="合同",
            content=(
                f"合同 {contract_no} 由 {payload.source or 'go'} 触发整单取消；"
                f"清理 factory_plan={stats.get('factory_plan', 0)}、units={stats.get('units', 0)}、"
                f"production_queue={stats.get('production_queue', 0)}、rush_order_queue={stats.get('rush_order_queue', 0)}；"
                f"沙盘重算：{'成功' if recompute_synced else '失败'}"
            ),
            user_id=operator,
            username=operator,
        )
        return {"status": "success", "cleanup": stats, "sandbox_recompute": recompute_synced}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Internal Contract Cancel Sync Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/orders/{order_id}")
def update_sales_order_api(
    order_id: str,
    payload: SalesOrderUpdatePayload,
    request: Request,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        df_orders = get_orders()
        mask = df_orders["订单号"].astype(str) == str(order_id)
        if not mask.any():
            raise HTTPException(status_code=404, detail="订单不存在")

        updates = payload.model_dump(exclude_unset=True)
        if "需求机型" in updates:
            _assert_models_in_dictionary(_extract_models_from_demand_text(str(updates.get("需求机型") or "")))
        for key, value in updates.items():
            if key == "需求数量" and value is not None:
                df_orders.loc[mask, key] = int(value)
            elif key == "发货时间":
                df_orders.loc[mask, key] = pd.to_datetime(value, errors="coerce") if value else None
            else:
                df_orders.loc[mask, key] = "" if value is None else str(value)
        save_orders(df_orders)
        changed_fields = [k for k, v in updates.items() if v is not None]
        append_audit_log(
            module="销售下单",
            action_type="修改",
            biz_type="订单",
            content=f"修改订单：{order_id}；更新字段：{', '.join(changed_fields) or '无'}",
            user_id=current_user.get("username"),
            username=current_operator,
        )

        # 【缺口2补全】订单被软删除时，清空沙盘中该订单关联卡片的合同字段，防止幽灵卡片并释放物理库存占用
        if updates.get("status") == "deleted":
            _clear_sandbox_units_by_order(str(order_id))
            inv_df = get_data()
            allocated_sns = inv_df[(inv_df["占用订单号"].astype(str) == str(order_id)) & (inv_df["状态"] != "已出库")]["流水号"].tolist()
            if allocated_sns:
                revert_to_inbound(allocated_sns, reason=f"订单软删除-自动解绑-{order_id}", operator=current_operator)

        return {"message": "订单更新成功"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新订单失败: {e}")


@router.delete("/orders/{order_id}")
def hard_delete_sales_order_api(
    order_id: str,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        from crud.orders import get_orders, save_orders
        df_orders = get_orders()
        mask = df_orders["订单号"].astype(str) == str(order_id)
        if not mask.any():
            raise HTTPException(status_code=404, detail="订单不存在")

        # 【缺口2补全】彻底删除前先清空沙盘关联卡片，防止幽灵卡片并释放物理库存占用
        _clear_sandbox_units_by_order(str(order_id))
        inv_df = get_data()
        allocated_sns = inv_df[(inv_df["占用订单号"].astype(str) == str(order_id)) & (inv_df["状态"] != "已出库")]["流水号"].tolist()
        if allocated_sns:
            revert_to_inbound(allocated_sns, reason=f"订单彻底删除-自动解绑-{order_id}", operator=current_operator)

        with get_engine().begin() as conn:
            conn.execute(text("DELETE FROM sales_orders WHERE `订单号` = :order_id"), {"order_id": str(order_id)})
        if hasattr(get_orders, "cache_clear"):
            get_orders.cache_clear()
        
        append_audit_log(
            module="销售下单",
            action_type="彻底删除",
            biz_type="订单",
            content=f"彻底删除订单：{order_id}",
            user_id=current_user.get("username"),
            username=current_operator,
        )
        return {"message": "订单已永久删除"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"彻底删除订单失败: {e}")


@router.get("/orders/{order_id}/allocations")
def get_order_allocations_api(order_id: str):
    try:
        order_id = str(order_id).strip()
        if not order_id:
            raise HTTPException(status_code=422, detail="订单号不能为空")
        rows = _get_order_contract_machine_rows(order_id)
        rows = rows.where(rows.notnull(), None)
        return {"data": rows.to_dict(orient="records")}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"读取配货记录失败: {e}")


@router.post("/orders/{order_id}/allocate")
def allocate_order_inventory_api(
    order_id: str,
    payload: OrderAllocatePayload,
    request: Request = None,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        selected = [str(x).strip() for x in (payload.selected_serial_nos or []) if str(x).strip()]
        if not selected:
            raise HTTPException(status_code=422, detail="请先选择要配货的机台")

        orders_df = get_orders()
        hit = orders_df[orders_df["订单号"].astype(str) == str(order_id)]
        if hit.empty:
            raise HTTPException(status_code=404, detail="订单不存在")
        first = hit.iloc[0]
        customer = str(first.get("客户名", "") or "")
        agent = str(first.get("代理商", "") or "")
        candidate_rows = _get_order_contract_machine_rows(str(order_id))
        valid_sns = set(
            candidate_rows[
                (candidate_rows["流水号"].astype(str).str.strip() != "")
                & (candidate_rows["状态"].astype(str).str.strip() != "已出库")
                & (
                    (candidate_rows["占用订单号"].astype(str).str.strip() == "")
                    | (candidate_rows["占用订单号"].astype(str).str.strip() == str(order_id))
                )
            ]["流水号"].astype(str).str.strip().tolist()
        )
        invalid = [sn for sn in selected if sn not in valid_sns]
        if invalid:
            raise HTTPException(status_code=422, detail=f"所选机台不属于该订单绑定合同，或已不可配货: {', '.join(invalid[:10])}")
        allocate_inventory(str(order_id), customer, agent, selected, operator=current_operator)
        append_audit_log(
            module="订单配货",
            action_type="配货",
            biz_type="订单",
            content=f"为订单 {order_id} 配货 {len(selected)} 台机台；流水号：{', '.join(selected[:10])}",
            user_id=_user_id_from_context(current_user),
            username=current_operator,
        )
        return {"message": f"配货成功，已锁定 {len(selected)} 台机台"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"配货失败: {e}")


@router.post("/orders/{order_id}/complete-allocation")
def complete_order_allocation_api(
    order_id: str,
    request: Request = None,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        order_id = str(order_id).strip()
        if not order_id:
            raise HTTPException(status_code=422, detail="订单号不能为空")

        orders_df = get_orders()
        hit = orders_df[orders_df["订单号"].astype(str).str.strip() == order_id]
        if hit.empty:
            raise HTTPException(status_code=404, detail="订单不存在")
        order_idx = hit.index[0]
        current_status = str(orders_df.at[order_idx, "status"] or "active")
        if current_status == "ready":
            return {"message": "配货已完成", "completed": True, "logged": 0}

        demand_counts = _parse_order_demand_counts(hit.iloc[0])
        if not demand_counts:
            raise HTTPException(status_code=422, detail="订单需求机型为空，无法完成配货")

        inv_df = get_data()
        if inv_df.empty:
            raise HTTPException(status_code=422, detail="配货未完成：未找到已配机台")
        for col in ["流水号", "机型", "状态", "占用订单号"]:
            if col not in inv_df.columns:
                inv_df[col] = ""

        allocated_df = inv_df[
            (inv_df["占用订单号"].astype(str).str.strip() == order_id)
            & (inv_df["状态"].astype(str).str.strip() != "已出库")
            & (inv_df["流水号"].astype(str).str.strip() != "")
        ].copy()
        allocated_counts: dict[str, int] = {}
        for _, row in allocated_df.iterrows():
            model = _normalize_alloc_model(row.get("机型", ""))
            if model:
                allocated_counts[model] = allocated_counts.get(model, 0) + 1

        missing: list[str] = []
        for model, need in demand_counts.items():
            allocated = allocated_counts.get(model, 0)
            if allocated < need:
                missing.append(f"{model} 缺少 {need - allocated} 台")
        if missing:
            raise HTTPException(status_code=422, detail=f"配货未完成：{'；'.join(missing)}")

        serials = allocated_df["流水号"].astype(str).str.strip().tolist()
        pending_inbound_df = allocated_df[allocated_df["状态"].astype(str).str.strip() == "待入库"].copy()
        pending_inbound_serials = pending_inbound_df["流水号"].astype(str).str.strip().tolist()
        pending_status_map = {
            str(row.get("流水号", "")).strip(): str(row.get("状态", "") or "").strip()
            for _, row in pending_inbound_df.iterrows()
        }
        if serials:
            from database import get_engine
            from sqlalchemy import text, bindparam
            from crud.inbound_history import record_inbound_history
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            try:
                with get_engine().begin() as conn:
                    conn.execute(
                        text("""
                            UPDATE finished_goods_data
                            SET 状态 = '待发货',
                                更新时间 = :now
                            WHERE 流水号 IN :sns AND 状态 != '已出库'
                        """).bindparams(bindparam("sns", expanding=True)),
                        {"now": now_str, "sns": serials}
                    )
                    if pending_inbound_serials:
                        record_inbound_history(
                            conn,
                            pending_inbound_serials,
                            source="配货自动入库",
                            operator=current_operator,
                            inbound_time=now_str,
                            status_before=pending_status_map,
                            status_after="待发货",
                        )
            except Exception as ex:
                raise HTTPException(status_code=500, detail=f"更新机台状态为待发货失败: {ex}")
            append_log("配货自动入库", serials, operator=current_operator)

        orders_df.at[order_idx, "status"] = "ready"
        save_orders(orders_df)
        append_audit_log(
            module="订单配货",
            action_type="配货完成",
            biz_type="订单",
            content=f"订单 {order_id} 配货完成，记录配货自动入库 {len(serials)} 台；流水号：{', '.join(serials[:10])}",
            user_id=_user_id_from_context(current_user),
            username=current_operator,
        )
        return {"message": "配货完成，订单已满足", "completed": True, "logged": len(serials)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"配货完成失败: {e}")


@router.post("/orders/{order_id}/release")
def release_order_inventory_api(
    order_id: str,
    payload: OrderReleasePayload,
    request: Request = None,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        order_id = str(order_id).strip()
        if not order_id:
            raise HTTPException(status_code=422, detail="订单号不能为空")

        inv_df = get_data()
        allocated_df = inv_df[
            (inv_df["占用订单号"].astype(str) == order_id)
            & (inv_df["状态"].astype(str) != "已出库")
        ]
        if allocated_df.empty:
            return {"message": "该订单当前没有可释放的配货机台", "released": 0}

        if payload.all:
            target_sns = allocated_df["流水号"].astype(str).tolist()
        else:
            selected = [str(x).strip() for x in (payload.selected_serial_nos or []) if str(x).strip()]
            if not selected:
                raise HTTPException(status_code=422, detail="请先选择要释放的机台")
            target_sns = allocated_df[allocated_df["流水号"].astype(str).isin(selected)]["流水号"].astype(str).tolist()
            if not target_sns:
                raise HTTPException(status_code=422, detail="所选机台不属于当前订单或已不可释放")

        revert_to_inbound(target_sns, reason=f"订单配货释放-{order_id}", operator=current_operator)
        orders_df = get_orders()
        hit = orders_df[orders_df["订单号"].astype(str).str.strip() == order_id]
        if not hit.empty:
            order_idx = hit.index[0]
            if str(orders_df.at[order_idx, "status"] or "active") == "ready":
                orders_df.at[order_idx, "status"] = "active"
                save_orders(orders_df)
        append_audit_log(
            module="订单配货",
            action_type="释放",
            biz_type="订单",
            content=f"释放订单 {order_id} 已配机台 {len(target_sns)} 台；流水号：{', '.join(target_sns[:10])}",
            user_id=_user_id_from_context(current_user),
            username=current_operator,
        )
        return {"message": f"已释放 {len(target_sns)} 台机台", "released": len(target_sns)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"释放配货失败: {e}")


@router.post("/contract/{contract_id}/status")
def update_contract_status(
    contract_id: str,
    payload: StatusPayload,
    request: Request,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        new_status = str(payload.status or "").strip()
        if not new_status:
            raise HTTPException(status_code=422, detail="status 不能为空")
        recompute_synced = False
        if new_status == "已规划":
            with get_engine().connect() as conn:
                exists = conn.execute(
                    text("SELECT 1 FROM factory_plan WHERE `合同号` = :cid LIMIT 1"),
                    {"cid": str(contract_id)},
                ).first()
            if not exists:
                raise HTTPException(status_code=404, detail="合同不存在")
            if not _trigger_sandbox_recompute_sync(current_user):
                raise HTTPException(status_code=502, detail="沙盘同步失败，合同状态未更新，请稍后重试")
            recompute_synced = True

        cleanup_stats: dict[str, int] = {}
        with get_engine().begin() as conn:
            if new_status == "已取消":
                cleanup_stats = _cleanup_cancelled_contract_links(conn, str(contract_id), current_operator)
                if cleanup_stats.get("factory_plan", 0) == 0:
                    raise HTTPException(status_code=404, detail="合同不存在")
            else:
                result = conn.execute(
                    text("UPDATE factory_plan SET `状态` = :status WHERE `合同号` = :cid"),
                    {"status": new_status, "cid": str(contract_id)},
                )
                if result.rowcount == 0:
                    raise HTTPException(status_code=404, detail="合同不存在")
                
        _clear_planning_related_caches()
        if new_status == "已取消":
            recompute_synced = _trigger_sandbox_recompute_sync(current_user)
        sync_text = "（已同步沙盘重算）" if recompute_synced else ""
        cancel_text = "（已同步清空沙盘所有状态批次卡片及排产队列，并触发预测沙盒重算）" if new_status == "已取消" else ""
        append_audit_log(
            module="合同管理",
            action_type="更新状态",
            biz_type="合同",
            content=f"合同 {contract_id} 状态更新为：{new_status}" + (cancel_text or sync_text),
            user_id=current_user.get("username"),
            username=current_operator,
        )
        return {
            "message": f"合同状态已更新为 {new_status}" + ("，对应沙盘卡片及队列已同步清理，预测沙盒已同步重算" if new_status == "已取消" else ("，沙盘已同步重算" if recompute_synced else "")),
            "sync": {"cleanup": cleanup_stats, "sandbox_recompute": recompute_synced},
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"状态更新失败: {e}")


@router.post("/contract/{contract_id}/link-order")
def link_contract_to_order(
    contract_id: str,
    payload: LinkOrderPayload,
    request: Request,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        order_id = str(payload.order_id or "").strip()
        if not order_id:
            raise HTTPException(status_code=422, detail="订单号不能为空")

        # 验证订单号是否存在
        orders_df = get_orders()
        if orders_df.empty or order_id not in orders_df["订单号"].values:
            raise HTTPException(status_code=404, detail=f"订单号 {order_id} 不存在")

        df_plan = get_factory_plan()
        mask = df_plan["合同号"].astype(str) == str(contract_id)
        if not mask.any():
            raise HTTPException(status_code=404, detail="合同不存在")

        existing = df_plan.loc[mask, "订单号"].fillna("").astype(str).str.strip()
        conflict = existing[(existing != "") & (existing != order_id)]
        if not conflict.empty:
            raise HTTPException(status_code=400, detail=f"合同已关联订单 {conflict.iloc[0]}，请先解除关联")

        _link_contracts_to_order([str(contract_id)], order_id)
        append_audit_log(
            module="合同管理",
            action_type="关联订单",
            biz_type="合同",
            content=f"合同 {contract_id} 关联订单：{order_id}",
            user_id=current_user.get("username"),
            username=current_operator,
        )

        return {"message": f"已成功将合同 {contract_id} 与订单 {order_id} 关联"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"关联订单失败: {e}")


def _process_contracts_batch(
    add_list: List[Dict[str, Any]],
    rush_source_rows: List[Dict[str, Any]],
    save_mode: str,
    is_rush: bool,
    user_ctx: dict,
    operator: str,
    background_tasks=None,
) -> dict:
    """Core contract creation logic, reusable from multiple endpoints."""
    if not add_list:
        raise HTTPException(status_code=422, detail="没有可新增记录（可能都已存在或字段不完整）")

    df_plan = get_factory_plan()
    now_status = "已规划" if save_mode == "spot" else "待规划"

    # Filter out duplicates within this batch call
    clean_add_list: List[Dict[str, Any]] = []
    clean_rush_rows: List[Dict[str, Any]] = []
    existed = 0
    for i, item in enumerate(add_list):
        cid = str(item["合同号"])
        model = str(item["机型"])
        dup_mask = (df_plan["合同号"].astype(str) == cid) & (df_plan["机型"].astype(str) == model)
        if dup_mask.any():
            existed += 1
            continue
        item["状态"] = now_status
        clean_add_list.append(item)
        if i < len(rush_source_rows):
            clean_rush_rows.append(rush_source_rows[i])

    if not clean_add_list:
        raise HTTPException(status_code=422, detail="没有可新增记录（可能都已存在或字段不完整）")

    df_plan = pd.concat([df_plan, pd.DataFrame(clean_add_list)], ignore_index=True)
    save_factory_plan(df_plan)

    if save_mode == "sandbox" and not is_rush and background_tasks:
        background_tasks.add_task(_trigger_sandbox_recompute_sync, user_ctx)

    rush_q_rows = clean_rush_rows if (is_rush and save_mode == "sandbox") else []
    rush_created = _insert_rush_order_queue(
        rush_q_rows,
        created_by=user_ctx.get("username") or operator,
    )
    rush_auto_inserted = (
        _auto_insert_rush_orders(rush_q_rows, user_ctx, operator)
        if RUSH_AUTO_INSERT_ON_ENTRY
        else 0
    )

    added_contract_ids = list(set([str(item["合同号"]) for item in clean_add_list]))
    contract_ids_str = "、".join(added_contract_ids)

    if save_mode == "spot" and added_contract_ids:
        with get_engine().begin() as conn:
            conn.execute(
                text(
                    "UPDATE rush_order_queue SET `status` = 'deleted', `updated_by` = :updated_by "
                    "WHERE TRIM(COALESCE(`contract_no`, '')) COLLATE utf8mb4_general_ci IN :cids "
                    "AND `status` = 'pending'"
                ).bindparams(bindparam("cids", expanding=True)),
                {
                    "cids": added_contract_ids,
                    "updated_by": str(user_ctx.get("username") or operator or ""),
                },
            )

    append_audit_log(
        module="合同管理",
        action_type="批量录入",
        biz_type="合同",
        content=f"批量录入合同 {len(clean_add_list)} 条（合同号：{contract_ids_str}）；跳过重复 {existed} 条；急单卡生成 {rush_created} 张；自动入沙盘 {rush_auto_inserted} 条",
        user_id=user_ctx.get("username"),
        username=operator,
    )
    return {
        "message": f"批量录入完成，新增 {len(clean_add_list)} 条，跳过重复 {existed} 条",
        "inserted": len(clean_add_list),
        "skipped": existed,
        "rush_created": rush_created,
        "rush_auto_inserted": rush_auto_inserted,
        "save_mode": save_mode,
        "contract_ids": added_contract_ids,
    }


@router.post("/contracts/batch-create")
def create_contracts_batch(
    payload: BatchContractCreatePayload,
    request: Request,
    background_tasks: BackgroundTasks,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        save_mode = str(payload.save_mode or "sandbox").strip().lower()
        if save_mode not in {"sandbox", "spot"}:
            raise HTTPException(status_code=422, detail="save_mode 仅支持 sandbox 或 spot")
        rows = payload.rows or []
        if not rows:
            raise HTTPException(status_code=422, detail="请至少提供 1 条合同记录")
        _assert_models_in_dictionary([str(item.机型 or "").strip() for item in rows])

        add_list: List[Dict[str, Any]] = []
        rush_source_rows: List[Dict[str, Any]] = []
        for item in rows:
            cid = str(item.合同号 or "").strip()
            customer = str(item.客户名 or "").strip()
            model = str(item.机型 or "").strip()
            due = str(item.要求交期 or "").strip()
            if not cid or not customer or not model or not due:
                continue
            qty = int(item.排产数量)
            add_list.append(
                {
                    "合同号": cid,
                    "机型": model,
                    "排产数量": qty,
                    "要求交期": due,
                    "状态": "",  # filled by _process_contracts_batch
                    "备注": str(item.备注 or "").strip(),
                    "客户名": customer,
                    "代理商": str(item.代理商 or "").strip(),
                    "指定批次/来源": {},
                    "订单号": "",
                }
            )
            rush_source_rows.append(
                {
                    "contract_no": cid,
                    "customer": customer,
                    "dealer_name": str(item.代理商 or "").strip(),
                    "model_type": model,
                    "due_date": due,
                    "qty": qty,
                    "remark": str(item.备注 or "").strip(),
                }
            )

        result = _process_contracts_batch(
            add_list=add_list,
            rush_source_rows=rush_source_rows,
            save_mode=save_mode,
            is_rush=bool(payload.is_rush),
            user_ctx=current_user,
            operator=current_operator,
            background_tasks=background_tasks,
        )

        # 如果来自经销商订单，回写 contract_no
        dealer_order_no = str(payload.dealer_order_no or "").strip()
        if dealer_order_no and result.get("contract_ids"):
            from crud.dealer_orders import mark_dealer_order_contracted
            try:
                contract_no_str = "、".join(result["contract_ids"])
                mark_dealer_order_contracted(
                    dealer_order_no,
                    contract_no=contract_no_str,
                    operator=current_operator,
                )
            except Exception as exc:
                import logging
                logging.getLogger(__name__).warning(
                    "Failed to update dealer_order %s after contract creation: %s",
                    dealer_order_no, exc,
                )

        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"批量录入失败: {e}")


@router.post("/contract/{contract_id}/edit-preview")
def preview_contract_edit(
    contract_id: str,
    payload: ContractEditPayload,
    request: Request,
):
    try:
        contract_id = str(contract_id or "").strip()
        if not contract_id:
            raise HTTPException(status_code=422, detail="合同号不能为空")
        if not payload.items:
            raise HTTPException(status_code=422, detail="至少保留一条机型明细")
        _assert_models_in_dictionary([str(item.机型 or "").strip() for item in payload.items])
        with get_engine().connect() as conn:
            _, _, _, preview = _load_contract_edit_context(conn, contract_id, payload, for_update=False)
        return preview
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"合同编辑预检失败: {e}")


@router.put("/contract/{contract_id}")
def edit_contract(
    contract_id: str,
    payload: ContractEditPayload,
    request: Request,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        contract_id = str(contract_id or "").strip()
        if not contract_id:
            raise HTTPException(status_code=422, detail="合同号不能为空")
        if not payload.items:
            raise HTTPException(status_code=422, detail="至少保留一条机型明细")
        _assert_models_in_dictionary([str(item.机型 or "").strip() for item in payload.items])

        sync_result: dict[str, Any] = {}
        conflicts: list[dict[str, Any]] = []
        with get_engine().begin() as conn:
            existing_rows, new_rows, order_id, preview = _load_contract_edit_context(conn, contract_id, payload, for_update=True)
            unit_plan = _validate_contract_edit_decision(preview, payload.mapping_decision)

            factory_stats = _upsert_factory_plan_contract_rows(conn, contract_id, new_rows)
            sales_order_rows = _sync_sales_order_from_contract_edit(conn, order_id, new_rows)
            queue_stats = _sync_production_queue_from_contract_edit(conn, contract_id, new_rows)
            rush_stats, rush_conflicts = _sync_rush_queue_from_contract_edit(conn, contract_id, new_rows, current_operator)
            unit_stats, unit_conflicts = _sync_units_from_contract_edit(conn, contract_id, new_rows, order_id, unit_plan)
            trace_stats, trace_conflicts = _sync_plan_import_and_inventory_from_contract_edit(conn, contract_id, new_rows, order_id)
            conflicts.extend(rush_conflicts)
            conflicts.extend(unit_conflicts)
            conflicts.extend(trace_conflicts)

            old_demand_text, old_total_qty = _contract_demand_text(existing_rows)
            new_demand_text, new_total_qty = _contract_demand_text(new_rows)
            sync_result = {
                "factory_plan": factory_stats,
                "sales_orders": {"updated": sales_order_rows},
                "production_queue": queue_stats,
                "rush_order_queue": rush_stats,
                "units": unit_stats,
                "trace_tables": trace_stats,
                "old_demand": {"text": old_demand_text, "quantity": old_total_qty},
                "new_demand": {"text": new_demand_text, "quantity": new_total_qty},
                "order_id": order_id,
                "impact_preview": preview,
            }

        _clear_planning_related_caches()
        recompute_synced = _trigger_sandbox_recompute_sync(current_user)
        sync_result["sandbox_recompute"] = recompute_synced

        append_audit_log(
            module="合同管理",
            action_type="编辑",
            biz_type="合同",
            content=(
                f"编辑合同：{contract_id}；"
                f"需求 {sync_result.get('old_demand', {}).get('text', '') or '-'}"
                f" → {sync_result.get('new_demand', {}).get('text', '') or '-'}；"
                f"沙盘重算：{'成功' if recompute_synced else '失败'}；"
                f"冲突 {len(conflicts)} 项"
            ),
            user_id=current_user.get("username"),
            username=current_operator,
        )
        message = "合同修改已保存，已同步全局数据"
        if conflicts:
            message += f"，有 {len(conflicts)} 项需要人工确认"
        if not recompute_synced:
            message += "，但沙盘重算触发失败，请稍后手动刷新沙盘"
        return {
            "message": message,
            "sync": sync_result,
            "conflicts": conflicts,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"合同编辑失败: {e}")


@router.get("/contract/{contract_id}/files")
def get_contract_files_api(contract_id: str):
    try:
        df = get_contract_files(contract_id)
        df = df.where(df.notnull(), None)
        return {"data": df.to_dict(orient="records")}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取附件失败: {e}")


@router.post("/contract/{contract_id}/files")
async def upload_contract_file_api(
    contract_id: str,
    file: UploadFile = File(...),
    customer_name: str = "",
    uploader_name: str = "",
    current_operator: str = Depends(get_current_operator_name),
):
    try:
        ok, msg = await asyncio.to_thread(
            save_contract_file,
            file,
            customer_name or str(contract_id),
            contract_id,
            uploader_name or current_operator or "API",
            True,
        )
        if not ok:
            raise HTTPException(status_code=422, detail=msg)
        return {"message": msg}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"上传附件失败: {e}")


@router.delete("/contract/{contract_id}/files/{file_name}")
def delete_contract_file_api(
    contract_id: str,
    file_name: str,
    current_operator: str = Depends(get_current_operator_name),
):
    try:
        decoded_name = unquote(file_name)
        ok, msg = delete_contract_file(contract_id, decoded_name, operator=current_operator or "API")
        if not ok:
            raise HTTPException(status_code=422, detail=msg)
        return {"message": msg}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除附件失败: {e}")


@router.get("/contract/{contract_id}/files/{file_name}/download")
def download_contract_file_api(contract_id: str, file_name: str):
    try:
        decoded_name = unquote(file_name)
        df = get_contract_files(contract_id)
        if df.empty:
            raise HTTPException(status_code=404, detail="附件不存在")
        hit = df[df["file_name"].astype(str) == decoded_name]
        if hit.empty:
            raise HTTPException(status_code=404, detail="附件不存在")
        rel_path = str(hit.iloc[0].get("file_path", "")).strip()
        if not rel_path:
            raise HTTPException(status_code=404, detail="附件路径无效")
        abs_path = os.path.join(BASE_DIR, rel_path)
        if not os.path.exists(abs_path):
            raise HTTPException(status_code=404, detail="附件文件不存在")
        return FileResponse(path=abs_path, filename=decoded_name)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"下载附件失败: {e}")


@router.get("/contract/{contract_id}/files/{file_name}/preview")
def preview_contract_file_api(contract_id: str, file_name: str):
    try:
        decoded_name = unquote(file_name)
        df = get_contract_files(contract_id)
        if df.empty:
            raise HTTPException(status_code=404, detail="附件不存在")
        hit = df[df["file_name"].astype(str) == decoded_name]
        if hit.empty:
            raise HTTPException(status_code=404, detail="附件不存在")

        rel_path = str(hit.iloc[0].get("file_path", "")).strip()
        if not rel_path:
            raise HTTPException(status_code=404, detail="附件路径无效")
        abs_path = os.path.join(BASE_DIR, rel_path)
        if not os.path.exists(abs_path):
            raise HTTPException(status_code=404, detail="附件文件不存在")

        ext = os.path.splitext(decoded_name)[1].lower()
        if ext in {".pdf", ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}:
            mime_map = {
                ".pdf": "application/pdf",
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
                ".png": "image/png",
                ".gif": "image/gif",
                ".bmp": "image/bmp",
                ".webp": "image/webp",
            }
            with open(abs_path, "rb") as f:
                encoded = base64.b64encode(f.read()).decode("ascii")
            return {
                "type": "url",
                "url": f"data:{mime_map[ext]};base64,{encoded}",
                "ext": ext,
            }

        if ext == ".docx":
            try:
                import mammoth
            except ImportError:
                raise HTTPException(status_code=422, detail="服务端缺少 mammoth，暂不支持 DOCX 在线预览")
            with open(abs_path, "rb") as f:
                result = mammoth.convert_to_html(f)
            return {
                "type": "html",
                "html": result.value or "",
                "ext": ext,
            }

        if ext == ".doc":
            return {
                "type": "legacy-doc",
                "ext": ext,
                "message": "DOC 为旧版 Word 二进制格式，当前仅支持下载后用 Word/WPS 查看；DOCX 可在线预览。",
            }

        return {"type": "", "ext": ext, "message": "该文件类型暂不支持在线预览"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"预览附件失败: {e}")


@router.post("/contract/{contract_id}/save-plan")
def save_contract_plan(
    contract_id: str,
    payload: PlanSavePayload,
    request: Request,
    current_operator: str = Depends(get_current_operator_name),
    current_user: dict = Depends(get_current_user_context),
):
    try:
        if not payload.rows:
            raise HTTPException(status_code=422, detail="缺少规划数据")

        # 1. 采用 SQL 精准 UPDATE，避免依赖 get_factory_plan 丢失主键及全表重写
        with get_engine().begin() as conn:
            # 验证合同是否存在
            res = conn.execute(
                text("SELECT 1 FROM factory_plan WHERE `合同号` = :cid LIMIT 1"),
                {"cid": str(contract_id)}
            )
            if not res.fetchone():
                raise HTTPException(status_code=404, detail="合同不存在")

            for row_payload in payload.rows:
                idx = int(row_payload.row_index)
                alloc = {}
                for k, v in (row_payload.allocation or {}).items():
                    qty = int(v or 0)
                    if qty > 0:
                        alloc[str(k)] = qty
                alloc_json = json.dumps(alloc, ensure_ascii=False)
                
                # 更新指定批次和状态
                if payload.mark_to_planned:
                    conn.execute(
                        text("""
                            UPDATE factory_plan
                            SET `指定批次/来源` = :alloc,
                                `状态` = CASE WHEN `状态` = '待规划' THEN '已规划' ELSE `状态` END
                            WHERE id = :id AND `合同号` = :cid
                        """),
                        {"alloc": alloc_json, "id": idx, "cid": str(contract_id)}
                    )
                else:
                    # 核心改动：如果未全部分配（mark_to_planned 为 False），
                    # 且数据库中该行状态已是"已规划"，则将其回退为"待规划"
                    conn.execute(
                        text("""
                            UPDATE factory_plan
                            SET `指定批次/来源` = :alloc,
                                `状态` = CASE WHEN `状态` = '已规划' THEN '待规划' ELSE `状态` END
                            WHERE id = :id AND `合同号` = :cid
                        """),
                        {"alloc": alloc_json, "id": idx, "cid": str(contract_id)}
                    )
        
        # 2. 清理缓存，确保下一次读取获取到最新数据
        from crud.planning import get_factory_plan
        get_factory_plan.cache_clear()

        # 3. 同步写回 sales_orders 的指定批次/来源（保持原有逻辑）
        df_plan = get_factory_plan()
        contract_rows = df_plan[df_plan["合同号"].astype(str) == str(contract_id)]
        order_id = str(contract_rows.iloc[0].get("订单号", "") or "").strip() if not contract_rows.empty else ""
        
        if order_id:
            all_plans: Dict[str, Dict[str, int]] = {}
            for _, row in contract_rows.iterrows():
                model_name = str(row.get("机型", "")).strip()
                alloc_data = row.get("指定批次/来源", {})
                if isinstance(alloc_data, str):
                    alloc_data = parse_alloc_dict(alloc_data)
                if model_name and alloc_data:
                    all_plans[model_name] = alloc_data
            if all_plans:
                orders_df = get_orders()
                hit = orders_df["订单号"].astype(str) == order_id
                if hit.any():
                    orders_df.loc[hit, "指定批次/来源"] = [all_plans] * int(hit.sum())
                    save_orders(orders_df)

        append_audit_log(
            module="合同管理",
            action_type="保存规划",
            biz_type="合同",
            content=f"保存合同 {contract_id} 的规划，共 {len(payload.rows)} 行",
            user_id=current_user.get("username"),
            username=current_operator,
        )
        return {"message": "规划保存成功"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"规划保存失败: {e}")


@router.get("/export-production-history")
def export_production_history(sheet: str = "all"):
    """导出截至目前的所有排产数据（包括生产中和已完工的机台明细）为 Excel 格式，包含排产台账和跟踪单"""
    from collections import OrderedDict, defaultdict
    try:
        include_ledger = sheet in ("all", "ledger")
        include_tracking = sheet in ("all", "tracking")

        with get_engine().connect() as conn:
            if include_ledger:
                sql_ledger = """
                    SELECT 
                        phl.unit_id AS `机台ID`,
                        COALESCE(phl.production_line_name, '待排产队列') AS `产线`,
                        phl.batch_code AS `批次号`,
                        phl.model_type AS `机型`,
                        phl.contract_no AS `合同号`,
                        phl.customer AS `客户名`,
                        phl.dealer_name AS `经销商`,
                        CASE phl.status
                            WHEN 'Completed' THEN '已完工'
                            WHEN 'In_Production' THEN '生产中'
                            WHEN 'Cancelled' THEN '已撤销'
                            ELSE phl.status
                        END AS `状态`,
                        COALESCE(fgb.batch_inbound_date, b.expected_inbound_date, fg.`预计入库时间`) AS `预计入库时间`,
                        COALESCE(u.is_locked, 0) AS `锁定状态`,
                        phl.order_remark AS `备注`,
                        phl.scheduled_at AS `排产上线时间`,
                        phl.completed_at AS `完工时间`,
                        COALESCE(u.serial_no, u.forecast_serial_no) AS `流水号`
                    FROM production_history_ledger phl
                    LEFT JOIN units u ON u.unit_id = phl.unit_id
                    LEFT JOIN batches b ON (b.batch_code COLLATE utf8mb4_general_ci = phl.batch_code COLLATE utf8mb4_general_ci AND phl.batch_code <> '') 
                        OR b.batch_id COLLATE utf8mb4_general_ci = SUBSTRING_INDEX(phl.unit_id, '-', 5) COLLATE utf8mb4_general_ci
                    LEFT JOIN (
                        SELECT `批次号` AS batch_code, MIN(`预计入库时间`) AS batch_inbound_date 
                        FROM finished_goods_data 
                        WHERE TRIM(COALESCE(`批次号`, '')) <> '' 
                        GROUP BY `批次号`
                    ) fgb ON TRIM(COALESCE(fgb.batch_code, '')) COLLATE utf8mb4_general_ci = TRIM(COALESCE(b.batch_code, '')) COLLATE utf8mb4_general_ci
                    LEFT JOIN finished_goods_data fg ON fg.`流水号` COLLATE utf8mb4_general_ci = COALESCE(u.serial_no, u.forecast_serial_no) COLLATE utf8mb4_general_ci
                    WHERE phl.status IN ('In_Production', 'Completed')
                    ORDER BY phl.scheduled_at DESC, phl.id DESC
                """
                df_ledger = pd.read_sql(text(sql_ledger), conn)
                records = df_ledger.to_dict('records')
            else:
                records = []

            if include_tracking:
                sql_ledger_for_t = """
                    SELECT 
                        phl.unit_id AS `机台ID`,
                        phl.batch_code AS `批次号`,
                        phl.model_type AS `机型`,
                        phl.order_remark AS `备注`,
                        COALESCE(u.serial_no, u.forecast_serial_no) AS `流水号`
                    FROM production_history_ledger phl
                    LEFT JOIN units u ON u.unit_id = phl.unit_id
                    LEFT JOIN batches b ON (b.batch_code COLLATE utf8mb4_general_ci = phl.batch_code COLLATE utf8mb4_general_ci AND phl.batch_code <> '') 
                        OR b.batch_id COLLATE utf8mb4_general_ci = SUBSTRING_INDEX(phl.unit_id, '-', 5) COLLATE utf8mb4_general_ci
                    WHERE phl.status IN ('In_Production', 'Completed')
                      AND b.status IN ('Confirmed', 'In_Production')
                """
                df_ledger_for_t = pd.read_sql(text(sql_ledger_for_t), conn)
                records_for_t = df_ledger_for_t.to_dict('records')

                sql_pending = """
                    SELECT 
                        u.unit_id AS `机台ID`,
                        '待排产队列' AS `产线`,
                        COALESCE(b.batch_code, '-') AS `批次号`,
                        u.model_type AS `机型`,
                        u.contract_no AS `合同号`,
                        u.customer AS `客户名`,
                        u.dealer_name AS `经销商`,
                        '待排产' AS `状态`,
                        b.expected_inbound_date AS `预计入库时间`,
                        COALESCE(u.is_locked, 0) AS `锁定状态`,
                        u.order_remark AS `备注`,
                        u.created_at AS `排产上线时间`,
                        NULL AS `完工时间`,
                        COALESCE(u.serial_no, u.forecast_serial_no) AS `流水号`
                    FROM units u
                    LEFT JOIN batches b ON b.batch_id = u.batch_id
                    WHERE u.status = 'Pending'
                      AND b.status IN ('Confirmed', 'In_Production')
                    ORDER BY u.created_at DESC, u.unit_id DESC
                """
                df_pending = pd.read_sql(text(sql_pending), conn)
                records_pending = df_pending.to_dict('records')
            else:
                records_for_t = []
                records_pending = []

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import TextBlock, CellRichText

        wb = Workbook()

        # 创建 Sheet
        if include_ledger and include_tracking:
            ws = wb.active
            ws.title = "排产台账"
            ws2 = wb.create_sheet(title="跟踪单")
        elif include_ledger:
            ws = wb.active
            ws.title = "排产台账"
        elif include_tracking:
            ws2 = wb.active
            ws2.title = "跟踪单"

        # 样式定义
        header_fill = PatternFill(start_color="5C765C", end_color="5C765C", fill_type="solid")
        header_font = Font(name="宋体", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        green_fill = PatternFill(start_color="EAF2E8", end_color="EAF2E8", fill_type="solid")
        orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        # 辅助宽度计算函数
        def get_display_width(s):
            width = 0
            for char in s:
                if ord(char) > 127:
                    width += 2
                else:
                    width += 1
            return width

        def format_cell_value(model_type, remark, qty, target_width=20):
            # Return tuple of (model_text, quantity) for rich text formatting
            if remark:
                model_text = f"{model_type} {remark}"
            else:
                model_text = model_type
            return (model_text, qty)

        # ------------------ 填充 Sheet 1: 排产台账 ------------------
        if include_ledger:
            # Optimize column layout: separate 7055 and 8055, remove expected inbound time
            # 300 (narrow), 400, 500, 600/8060 (merged), 7055, 8055
            model_columns = ["300", "400", "500", "600", "7055", "8055"]

            batches = OrderedDict()
            for r in records:
                batch_code = str(r.get("批次号") or "").strip()
                if not batch_code:
                    batch_code = "-"
                if batch_code not in batches:
                    batches[batch_code] = {
                        "batch_code": batch_code,
                        "units": [],
                        "due_dates": set()
                    }
                batches[batch_code]["units"].append(r)
                due_date = r.get("预计入库时间")
                if due_date and str(due_date).strip() not in ("", None, "-", "None", "NaT"):
                    batches[batch_code]["due_dates"].add(due_date)

            ws.views.sheetView[0].showGridLines = True
            headers = ["批次号"] + model_columns + ["合计"]
            ws.append(headers)

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            start_row = 2
            total_cols = len(headers)

            sorted_batches = sorted(batches.items(), key=lambda x: (x[0] == "-", x[0]))
            batch_index = 0  # Track batch index for alternating colors
            for batch_code, batch in sorted_batches:
                # Determine fill color based on batch index (alternating)
                if batch_index % 2 == 0:
                    batch_fill = green_fill
                else:
                    batch_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
                batch_index += 1

                unique_pairs = []
                pair_qty = defaultdict(int)
                for unit in batch["units"]:
                    orig_model = str(unit.get("机型") or "").strip()
                    remark = str(unit.get("备注") or "").strip()
                    if remark in ("None", "none", "null", "NULL"):
                        remark = ""
                    pair = (orig_model, remark)
                    if pair_qty[pair] == 0:
                        unique_pairs.append(pair)
                    pair_qty[pair] += 1

                matched_by_col = defaultdict(list)
                special_pairs = []
                for pair in unique_pairs:
                    orig_model, remark = pair
                    combined = (orig_model + remark).upper()
                    matched_col = None

                    # Match models to columns with merging logic
                    # Check longer patterns first to avoid substring matching issues
                    if "8060" in combined:
                        matched_col = "600"  # Merge 8060 with 600
                    elif "8055" in combined:
                        matched_col = "8055"
                    elif "7055" in combined:
                        matched_col = "7055"
                    elif "600" in combined:
                        matched_col = "600"
                    elif "500" in combined:
                        matched_col = "500"
                    elif "400" in combined:
                        matched_col = "400"
                    elif "300" in combined:
                        matched_col = "300"

                    if matched_col:
                        matched_by_col[matched_col].append(pair)
                    else:
                        special_pairs.append(pair)

                col_assigned_count = {col: len(matched_by_col[col]) for col in model_columns}
                for pair in special_pairs:
                    best_col = min(model_columns, key=lambda c: col_assigned_count[c])
                    matched_by_col[best_col].append(pair)
                    col_assigned_count[best_col] += 1

                entries_by_model = {}
                for col_key in model_columns:
                    entries = []
                    for orig_model, remark in matched_by_col[col_key]:
                        qty = pair_qty[(orig_model, remark)]
                        entries.append(format_cell_value(orig_model, remark, qty, target_width=20))
                    entries_by_model[col_key] = entries

                max_rows = max(len(entries) for entries in entries_by_model.values())
                if max_rows == 0:
                    max_rows = 1

                end_row = start_row + max_rows - 1

                for r in range(start_row, end_row + 1):
                    for c in range(1, total_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.font = Font(name="宋体", size=10)
                        cell.border = thin_border
                        if c == total_cols:
                            cell.fill = orange_fill
                        else:
                            cell.fill = batch_fill  # Use alternating batch color
                        if c == 1 or c == total_cols - 1 or c == total_cols:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                ws.cell(row=start_row, column=1).value = batch_code
                if end_row > start_row:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

                for m_idx, m_type in enumerate(model_columns, start=2):
                    entries = entries_by_model[m_type]
                    for i, entry in enumerate(entries):
                        cell = ws.cell(row=start_row + i, column=m_idx)

                        # entry is a tuple of (model_text, qty)
                        if entry:
                            model_text, qty = entry
                            content_length = len(model_text) + len(str(qty))

                            # Determine font size based on content length
                            if content_length > 100:
                                font_size = 7
                            elif content_length > 60:
                                font_size = 8
                            elif content_length > 40:
                                font_size = 9
                            else:
                                font_size = 10

                            # Always use rich text with colored quantity
                            rich_text = CellRichText(
                                TextBlock(InlineFont(rFont="宋体", sz=font_size, color="000000"), f"{model_text}  "),
                                TextBlock(InlineFont(rFont="宋体", sz=font_size, b=True, color="FF0000"), str(qty))
                            )
                            cell.value = rich_text

                    if len(entries) <= 1 and end_row > start_row:
                        ws.merge_cells(start_row=start_row, start_column=m_idx, end_row=end_row, end_column=m_idx)

                total_qty = len(batch["units"])
                ws.cell(row=start_row, column=total_cols).value = total_qty
                if end_row > start_row:
                    ws.merge_cells(start_row=start_row, start_column=total_cols, end_row=end_row, end_column=total_cols)

                start_row = end_row + 1

            # Adjust column widths for A4 landscape - now 8 columns total (removed 预计入库时间)
            # Columns: 批次号, 300, 400, 500, 600/8060, 7055, 8055, 合计
            ws.column_dimensions['A'].width = 10  # 批次号
            ws.column_dimensions['B'].width = 14  # 300 (narrower)
            for col_idx in range(3, 8):  # 400, 500, 600/8060, 7055, 8055
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 20  # Model columns
            ws.column_dimensions[get_column_letter(8)].width = 8   # 合计
            # Total: 10 + 14 + 5*20 + 8 = 132 units (fits A4 landscape well)

            # Set page setup for A4 landscape - allow content to flow naturally
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToWidth = 1  # Fit columns to one page width
            ws.page_setup.fitToHeight = 0  # Allow unlimited pages vertically
            ws.print_options.horizontalCentered = True

            # Disable auto-scaling to prevent content compression
            ws.page_setup.scale = 100  # Use 100% scale, no shrinking
            ws.sheet_properties.pageSetUpPr.fitToPage = False  # Disable fit-to-page

            # Set print area to ensure all columns are included
            ws.print_area = f'A1:H{ws.max_row}'

        # ------------------ 填充 Sheet 2: 跟踪单 ------------------
        if include_tracking:
            ws2.views.sheetView[0].showGridLines = True
            headers2 = ["生产批次", "型号", "生产编号"]
            ws2.append(headers2)

            for col_idx in range(1, len(headers2) + 1):
                cell = ws2.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            combined_records = records_for_t + records_pending

            batches_tracking = defaultdict(list)
            for r in combined_records:
                b_code = str(r.get("批次号") or "").strip()
                if not b_code:
                    b_code = "-"
                batches_tracking[b_code].append(r)

            sorted_tracking_batches = sorted(batches_tracking.items(), key=lambda x: (x[0] == "-", x[0]))

            start_row_t = 2
            for b_code, b_units in sorted_tracking_batches:
                def sort_key(u):
                    sn = str(u.get("流水号") or "").strip()
                    uid = str(u.get("机台ID") or "").strip()
                    return (sn == "", sn, uid)

                sorted_units = sorted(b_units, key=sort_key)
                num_units = len(sorted_units)
                end_row_t = start_row_t + num_units - 1

                ws2.cell(row=start_row_t, column=1).value = b_code
                if end_row_t > start_row_t:
                    ws2.merge_cells(start_row=start_row_t, start_column=1, end_row=end_row_t, end_column=1)

                for i, u in enumerate(sorted_units):
                    row_idx = start_row_t + i
                    model = str(u.get("机型") or "").strip()
                    remark = str(u.get("备注") or "").strip()
                    if remark in ("None", "none", "null", "NULL"):
                        remark = ""

                    model_text = f"{model} {remark}" if remark else model
                    sn_text = str(u.get("流水号") or u.get("机台ID") or "").strip() or "-"

                    ws2.cell(row=row_idx, column=2).value = model_text
                    ws2.cell(row=row_idx, column=3).value = sn_text

                    for c in range(1, 4):
                        cell = ws2.cell(row=row_idx, column=c)
                        cell.font = Font(name="宋体", size=10)
                        cell.border = thin_border
                        cell.fill = green_fill
                        if c == 1 or c == 3:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                start_row_t = end_row_t + 1

            for col in ws2.columns:
                max_len = 0
                for cell in col:
                    val = str(cell.value or '')
                    w = get_display_width(val)
                    if w > max_len:
                        max_len = w
                col_letter = get_column_letter(col[0].column)
                ws2.column_dimensions[col_letter].width = max(max_len + 4, 12)

        import io
        from fastapi.responses import StreamingResponse

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_prefix = "production_history" if sheet != "tracking" else "production_tracking"
        headers = {
            "Content-Disposition": f"attachment; filename={file_prefix}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        }
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出排产数据失败: {e}")


class ExportExcelPayload(BaseModel):
    filename: str
    sheet_name: str
    headers: List[str]
    rows: List[List[Any]]


@router.post("/export-excel")
def export_excel(payload: ExportExcelPayload):
    """通用数据导出为 Excel xlsx 格式"""
    try:
        import io
        from fastapi.responses import StreamingResponse
        
        df = pd.DataFrame(payload.rows, columns=payload.headers)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=payload.sheet_name)
        output.seek(0)
        
        headers = {
            "Content-Disposition": f"attachment; filename={payload.filename}.xlsx"
        }
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出 Excel 失败: {e}")

